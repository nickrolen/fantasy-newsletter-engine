#!/usr/bin/env python3
"""
format_stats_report.py

Converts stats_report_weekN.json into newsletter-ready markdown organized by
newsletter section. Eliminates the LLM extraction step by doing all data
assembly in Python.

Usage:
    python format_stats_report.py --week 15
    python format_stats_report.py --input output/stats_report_week15.json
    python format_stats_report.py --week 15 --output output/stats_report_week15.md
    python format_stats_report.py --week 15 --last-week-recap config/LAST_WEEK_RECAP.md
"""

import argparse
import json
import re
import sys
from pathlib import Path
from datetime import datetime
from typing import Any, Optional

# Add project root to path for config imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from modules.data_loader import MANAGERS, CURRENT_SEASON

try:
    import openpyxl  # type: ignore
except ImportError:  # optional dependency
    openpyxl = None


# =============================================================================
# LAST WEEK BETTING PREDICTIONS (from previous week's JSON)
# =============================================================================

def load_last_week_betting_predictions(current_week: int, output_dir: Path) -> dict:
    """
    Load the previous week's stats report and extract betting predictions.
    
    Args:
        current_week: The current week number
        output_dir: Path to the output/ directory containing stats_report_weekN.json files
    
    Returns dict mapping matchup tuple to prediction info:
    {
        ("Benton", "Nick"): {
            "favorite": "Nick",
            "underdog": "Benton", 
            "spread": -36.5,
            "win_prob_favorite": 55.52,
            "win_prob_underdog": 44.48,
        },
    }
    """
    predictions = {}
    
    if current_week <= 1:
        return predictions
    
    last_week = current_week - 1
    last_week_path = output_dir / f"stats_report_week{last_week}.json"
    
    if not last_week_path.exists():
        return predictions
    
    try:
        with open(last_week_path, encoding="utf-8") as f:
            last_data = json.load(f)
    except Exception:
        return predictions
    
    # Extract betting predictions from looking_ahead
    # NOTE: looking_ahead is None when the generator ran with --fast (skipped sims),
    # so default-dict idiom is unsafe -- use `or {}`.
    looking_ahead = last_data.get("looking_ahead") or {}
    matchup_previews = looking_ahead.get("matchup_previews") or []
    
    for preview in matchup_previews:
        betting = preview.get("betting_line", {})
        if not betting:
            continue
        
        mgr_a = betting.get("manager_a", preview.get("manager_a"))
        mgr_b = betting.get("manager_b", preview.get("manager_b"))
        win_prob_a = betting.get("win_prob_a", 50)
        win_prob_b = betting.get("win_prob_b", 50)
        spread_a = betting.get("spread_a", 0)
        spread_b = betting.get("spread_b", 0)
        
        # Determine favorite (higher win probability)
        if win_prob_a >= win_prob_b:
            favorite = mgr_a
            underdog = mgr_b
            fav_prob = win_prob_a
            dog_prob = win_prob_b
            spread = spread_a  # Negative for favorite
        else:
            favorite = mgr_b
            underdog = mgr_a
            fav_prob = win_prob_b
            dog_prob = win_prob_a
            spread = spread_b
        
        # Store with sorted tuple key for consistent lookup
        key = tuple(sorted([mgr_a, mgr_b]))
        predictions[key] = {
            "favorite": favorite,
            "underdog": underdog,
            "spread": spread,
            "win_prob_favorite": fav_prob,
            "win_prob_underdog": dog_prob,
            "mgr_a": mgr_a,
            "mgr_b": mgr_b,
        }
    
    return predictions


def load_last_week_grades(current_week: int, output_dir: Path) -> dict:
    """
    Load the previous week's report card grades.
    
    Returns dict mapping manager name to letter grade:
    {"Nick": "B", "Benton": "C+", ...}
    """
    if current_week <= 1:
        return {}
    
    last_week_path = output_dir / f"stats_report_week{current_week - 1}.json"
    
    if not last_week_path.exists():
        return {}
    
    try:
        with open(last_week_path, encoding="utf-8") as f:
            last_data = json.load(f)
    except Exception:
        return {}
    
    return {
        rc["manager"]: rc.get("letter_grade", "?")
        for rc in last_data.get("report_cards", [])
    }


def load_last_week_title_odds(current_week: int, output_dir: Path) -> dict:
    """
    Load the previous week's title odds from power_rankings.
    
    Returns dict mapping manager name to title odds percentage:
    {"Nick": 74.85, "Benton": 25.15, "Garrett": 0.0, "Hayden": 0.0}
    """
    if current_week <= 1:
        return {}
    
    last_week_path = output_dir / f"stats_report_week{current_week - 1}.json"
    
    if not last_week_path.exists():
        return {}
    
    try:
        with open(last_week_path, encoding="utf-8") as f:
            last_data = json.load(f)
    except Exception:
        return {}
    
    # power_rankings is None when the prior week ran with --fast (no sims) -- use `or []`.
    return {
        pr["manager"]: pr.get("title_odds", 0.0)
        for pr in (last_data.get("power_rankings") or [])
    }


def _resolve_potw_path(base_path: Path) -> Path:
    """Find POTW_HISTORY.json -> checks config/ first, then base_path."""
    potw_path = base_path / "config" / "POTW_HISTORY.json"
    if potw_path.exists():
        return potw_path
    potw_path = base_path / "POTW_HISTORY.json"
    return potw_path


def _normalize_season_key(raw: str) -> str:
    """
    Normalize season key to 'YYYY-YY' format (NBA convention).
    
    '2025-2026' -> '2025-26'
    '2025-26' -> '2025-26' (already correct)
    """
    if not raw:
        return raw
    parts = raw.split("-")
    if len(parts) == 2 and len(parts[1]) == 4:
        return f"{parts[0]}-{parts[1][2:]}"
    return raw


def load_potw_history(current_week: int, base_path: Path = None, season: str = None) -> tuple[list[dict], dict]:
    """
    Load Player of the Week history from POTW_HISTORY.json.
    
    Args:
        current_week: Current week number (only loads weeks before this)
        base_path: Project base path to find POTW_HISTORY.json
        season: Current season key (e.g. '2025-26')
    
    Returns:
        Tuple of:
        - current_season_history: List of previous weeks' winners for this season
        - all_time: Full file data (all seasons) for career stats
    """
    if not base_path:
        return [], {}
    
    potw_path = _resolve_potw_path(base_path)
    if not potw_path.exists():
        return [], {}
    
    try:
        with open(potw_path, encoding="utf-8") as f:
            all_data = json.load(f)
    except Exception:
        return [], {}
    
    seasons = all_data.get("seasons", {})
    
    # Current season history (only weeks before current)
    season_key = _normalize_season_key(season) if season else None
    current_season = []
    if season_key and season_key in seasons:
        for entry in seasons[season_key]:
            week_num = entry.get("week", 0)
            if 0 < week_num < current_week:
                current_season.append({
                    "week": week_num,
                    "player": entry["player"],
                    "manager": entry.get("manager", "?"),
                    "total_fp": entry.get("total_fp", 0),
                    "games": entry.get("games", 0),
                })
        current_season.sort(key=lambda x: x["week"])
    
    return current_season, all_data


def save_potw_winner(base_path: Path, season: str, winner: dict, week: int) -> bool:
    """
    Append the current week's POTW winner to POTW_HISTORY.json.
    
    Skips if this week already has an entry. Writes back to disk.
    
    Args:
        base_path: Project base path
        season: Season key from metadata (e.g. '2025-2026' or '2025-26')
        winner: The player_of_week.winner dict from the stats report JSON
        week: Current week number
    
    Returns:
        True if written, False if skipped or error
    """
    if not base_path or not winner or not winner.get("player_name"):
        return False
    
    potw_path = _resolve_potw_path(base_path)
    
    # Load existing data (or start fresh)
    if potw_path.exists():
        try:
            with open(potw_path, encoding="utf-8") as f:
                all_data = json.load(f)
        except Exception:
            all_data = {}
    else:
        all_data = {
            "_comment": "All-time Player of the Week history. The formatter appends the current week's winner after each run. Organized by NBA season.",
        }
    
    if "seasons" not in all_data:
        all_data["seasons"] = {}
    
    season_key = _normalize_season_key(season)
    if season_key not in all_data["seasons"]:
        all_data["seasons"][season_key] = []
    
    season_list = all_data["seasons"][season_key]
    
    # Check if this week already has an entry
    existing_weeks = {e.get("week") for e in season_list}
    if week in existing_weeks:
        return False
    
    # Build entry with all available data
    entry = {
        "week": week,
        "player": winner["player_name"],
        "manager": winner.get("manager", "?"),
    }
    # Include rich data when available
    if winner.get("total_fp"):
        entry["total_fp"] = round(winner["total_fp"], 2)
    if winner.get("games"):
        entry["games"] = winner["games"]
    if winner.get("fppg"):
        entry["fppg"] = round(winner["fppg"], 2)
    
    season_list.append(entry)
    season_list.sort(key=lambda x: x.get("week", 0))
    
    # Write back
    try:
        potw_path.parent.mkdir(parents=True, exist_ok=True)
        with open(potw_path, "w", encoding="utf-8") as f:
            json.dump(all_data, f, indent=4, ensure_ascii=False)
        return True
    except Exception:
        return False


def compute_potw_career_stats(all_data: dict) -> dict:
    """
    Compute career POTW stats across all seasons.
    
    Returns dict like:
    {
        "by_player": {"Nikola Jokic": {"total": 5, "seasons": {"2025-26": 3, "2026-27": 2}}, ...},
        "by_manager": {"Benton": {"total": 8, "seasons": {"2025-26": 5, "2026-27": 3}}, ...},
    }
    """
    by_player = {}
    by_manager = {}
    
    for season_key, entries in all_data.get("seasons", {}).items():
        for entry in entries:
            player = entry.get("player", "?")
            manager = entry.get("manager", "?")
            
            if player not in by_player:
                by_player[player] = {"total": 0, "seasons": {}}
            by_player[player]["total"] += 1
            by_player[player]["seasons"][season_key] = by_player[player]["seasons"].get(season_key, 0) + 1
            
            if manager not in by_manager:
                by_manager[manager] = {"total": 0, "seasons": {}}
            by_manager[manager]["total"] += 1
            by_manager[manager]["seasons"][season_key] = by_manager[manager]["seasons"].get(season_key, 0) + 1
    
    return {"by_player": by_player, "by_manager": by_manager}


def load_player_projections(base_path: Path) -> dict:
    """
    Load projected FPPG for all players from PLAYERLIST.xlsx.

    Returns dict mapping player name to proj FPPG:
    {"Nikola Jokic": 59.77, "Luka Doncic": 52.85, ...}

    Notes:
    - Projections are expected to exist for normal newsletter generation.
    - This returns {} if the projections workbook can't be loaded; main() decides whether to hard-fail.
    """
    if openpyxl is None:
        return {}

    playerlist_path = base_path / "data" / "PLAYERLIST.xlsx"

    if not playerlist_path.exists():
        return {}

    wb = None
    try:
        wb = openpyxl.load_workbook(playerlist_path, read_only=True, data_only=True)
        ws = wb.active

        # Find column indices from header row
        headers = [cell.value for cell in ws[1]]
        name_col = headers.index("player_name")
        proj_col = headers.index("projectedFPPG")

        projections = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[name_col]
            proj = row[proj_col]
            # Keep zero-valued projections: an all-zero table is the legitimate
            # final-week state (no games left to project), and discarding zeros
            # would make a readable file look empty/unreadable to the caller.
            if name is not None and proj is not None:
                projections[name] = float(proj)

        return projections
    except Exception:
        return {}
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def load_injury_timelines(base_path: Path, current_week: int) -> dict:
    """
    Load injury timelines from INJURY_OVERRIDES.json.
    
    Returns dict mapping player name to timeline info:
    {
        "Anthony Davis": {
            "out_weeks": [12, 13, 14, 15, 16, 17, 18],
            "weeks_missed": 4,           # out_weeks <= current_week
            "total_projected": 7,         # len(out_weeks)
            "first_week_out": 12,
            "notes": "Hand injury - out ~7 weeks starting week 12",
            "return_week": None,
        }, ...
    }
    """
    # Check config/ first, then base_path directly
    io_path = base_path / "config" / "INJURY_OVERRIDES.json"
    if not io_path.exists():
        io_path = base_path / "INJURY_OVERRIDES.json"
    
    if not io_path.exists():
        return {}
    
    try:
        with open(io_path, encoding="utf-8") as f:
            io_data = json.load(f)
    except Exception:
        return {}
    
    timelines = {}
    for p in io_data.get("players", []):
        name = p.get("player_name", "")
        out_weeks = p.get("out_weeks", [])
        if not name or not out_weeks:
            continue
        
        weeks_missed = len([w for w in out_weeks if w <= current_week])
        
        timelines[name] = {
            "out_weeks": out_weeks,
            "weeks_missed": weeks_missed,
            "total_projected": len(out_weeks),
            "first_week_out": min(out_weeks),
            "notes": p.get("notes", ""),
            "return_week": p.get("return_week"),
        }
    
    return timelines


def load_weekly_context_notes(base_path: Path, current_week: int) -> str:
    """
    Load the 'notes' field from weeklycontextinput_weekN.json.
    Returns the notes string, or empty string if not found.
    """
    candidates = [
        base_path / f"weeklycontextinput_week{current_week}.json",
        base_path / "data" / f"weeklycontextinput_week{current_week}.json",
        base_path / "config" / f"weeklycontextinput_week{current_week}.json",
    ]
    for c in candidates:
        if c.exists():
            try:
                with open(c, encoding="utf-8") as f:
                    ctx = json.load(f)
                notes = ctx.get("notes", "")
                if isinstance(notes, list):
                    return " ".join(notes)
                return str(notes) if notes else ""
            except Exception:
                return ""
    return ""


def load_weekly_trades(base_path: Path, current_week: int, player_projs: dict, keeper_watch_players: list = None) -> list[dict]:
    """
    Load trades from the weekly context input file and enrich with player projections
    and keepability scores.
    
    Looks for weeklycontextinput_week{N}.json in base_path (or scripts/ subdirectory).
    
    Args:
        base_path: Project root path
        current_week: Current fantasy week number
        player_projs: Dict mapping player_name -> projected FPPG
        keeper_watch_players: Optional list of keeper_watch player dicts
            (from stats report JSON) with keepability_score and keeper_tier.
    
    Returns list of enriched trade dicts:
    [
        {
            "manager_a": "Hayden",
            "sends_a": [
                {"item": "Michael Porter Jr.", "is_player": True, "proj_fppg": 37.80,
                 "keepability_score": 52.3, "keeper_tier": "Strong Hold"},
                {"item": "2027 2nd round pick", "is_player": False}
            ],
            "manager_b": "Benton",
            "sends_b": [
                {"item": "Lauri Markkanen", "is_player": True, "proj_fppg": 37.52,
                 "keepability_score": 48.1, "keeper_tier": "On the Bubble"},
                {"item": "2026 5th round pick", "is_player": False},
                {"item": "2027 5th round pick", "is_player": False}
            ]
        }
    ]
    """
    # Build keepability lookup
    keeper_lookup = {}
    if keeper_watch_players:
        for p in keeper_watch_players:
            keeper_lookup[p["player_name"]] = {
                "keepability_score": p.get("keepability_score", 0),
                "keeper_tier": p.get("keeper_tier", ""),
            }
    # Search for weekly context file in common locations
    candidates = [
        base_path / f"weeklycontextinput_week{current_week}.json",
        base_path / "data" / f"weeklycontextinput_week{current_week}.json",
        base_path / "scripts" / f"weeklycontextinput_week{current_week}.json",
        base_path / "config" / f"weeklycontextinput_week{current_week}.json",
    ]
    
    context_path = None
    for c in candidates:
        if c.exists():
            context_path = c
            break
    
    if not context_path:
        return []
    
    try:
        with open(context_path, encoding="utf-8") as f:
            context = json.load(f)
    except Exception:
        return []
    
    raw_trades = context.get("trades", [])
    if not raw_trades:
        return []
    
    enriched = []
    for trade in raw_trades:
        enriched_trade = {
            "manager_a": trade.get("manager_a", "?"),
            "manager_b": trade.get("manager_b", "?"),
            "sends_a": [],
            "sends_b": [],
        }
        
        for side in ["sends_a", "sends_b"]:
            for item in trade.get(side, []):
                proj = player_projs.get(item)
                if proj is not None:
                    kd = keeper_lookup.get(item, {})
                    enriched_trade[side].append({
                        "item": item,
                        "is_player": True,
                        "proj_fppg": proj,
                        "keepability_score": kd.get("keepability_score", 0),
                        "keeper_tier": kd.get("keeper_tier", ""),
                    })
                else:
                    enriched_trade[side].append({
                        "item": item,
                        "is_player": False,
                    })
        
        enriched.append(enriched_trade)
    
    return enriched


def sync_weekly_trades_to_trades_json(
    base_path: Path, current_week: int, weekly_context_path: Optional[Path] = None
) -> int:
    """
    Auto-sync trades from weeklycontextinput_weekN.json into TRADES.json.

    For each trade in the weekly context input:
    1. Checks if a trade already exists in TRADES.json for the same week + managers
    2. If new, appends it to the trades array with the next trade_id
    3. Parses draft pick strings and updates draft_pick_ownership
    4. Saves TRADES.json

    Returns the number of new trades added (0 if all already existed).
    """
    # Find weekly context file
    if weekly_context_path is None:
        candidates = [
            base_path / f"weeklycontextinput_week{current_week}.json",
            base_path / "data" / f"weeklycontextinput_week{current_week}.json",
            base_path / "config" / f"weeklycontextinput_week{current_week}.json",
        ]
        for c in candidates:
            if c.exists():
                weekly_context_path = c
                break

    if weekly_context_path is None or not weekly_context_path.exists():
        return 0

    try:
        with open(weekly_context_path, encoding="utf-8") as f:
            context = json.load(f)
    except Exception:
        return 0

    raw_trades = context.get("trades", [])
    if not raw_trades:
        return 0

    # Find TRADES.json
    trades_path = base_path / "config" / "TRADES.json"
    if not trades_path.exists():
        trades_path = base_path / "TRADES.json"
    if not trades_path.exists():
        return 0

    try:
        with open(trades_path, encoding="utf-8") as f:
            trades_data = json.load(f)
    except Exception:
        return 0

    existing_trades = trades_data.get("trades", [])
    ownership = trades_data.get("draft_pick_ownership", {})

    # Find max trade_id for incrementing
    max_id = max((t.get("trade_id", 0) for t in existing_trades), default=0)

    # Get date range from context for the trade date
    date_range = context.get("date_range", "")

    added = 0
    for raw_trade in raw_trades:
        mgr_a = raw_trade.get("manager_a", "")
        mgr_b = raw_trade.get("manager_b", "")

        # Check for duplicate: same week + same pair of managers
        already_exists = False
        for existing in existing_trades:
            if existing.get("week") != current_week:
                continue
            e_a = existing.get("side_a", {}).get("manager", "")
            e_b = existing.get("side_b", {}).get("manager", "")
            if set([mgr_a, mgr_b]) == set([e_a, e_b]):
                already_exists = True
                break

        if already_exists:
            continue

        # Parse sends into players and picks
        sends_a_items = raw_trade.get("sends_a", [])
        sends_b_items = raw_trade.get("sends_b", [])

        sends_a_players = []
        sends_a_picks = []
        for item in sends_a_items:
            parsed_pick = _parse_pick_string(item)
            if parsed_pick:
                # Normalize to short format: "YYYY Nth (Manager)"
                year, rnd = parsed_pick
                ordinal = _ordinal(rnd)
                sends_a_picks.append(f"{year} {ordinal} ({mgr_a})")
            else:
                sends_a_players.append(item)

        sends_b_players = []
        sends_b_picks = []
        for item in sends_b_items:
            parsed_pick = _parse_pick_string(item)
            if parsed_pick:
                year, rnd = parsed_pick
                ordinal = _ordinal(rnd)
                sends_b_picks.append(f"{year} {ordinal} ({mgr_b})")
            else:
                sends_b_players.append(item)

        # Build the trade entry
        max_id += 1
        new_trade = {
            "trade_id": max_id,
            "week": current_week,
            "date": date_range,
            "side_a": {
                "manager": mgr_a,
                "sent_players": sends_a_players,
                "sent_picks": sends_a_picks,
            },
            "side_b": {
                "manager": mgr_b,
                "sent_players": sends_b_players,
                "sent_picks": sends_b_picks,
            },
        }
        existing_trades.append(new_trade)

        # Update draft pick ownership
        # Picks sent by A go to B, picks sent by B go to A
        for item in sends_a_items:
            _update_pick_ownership(ownership, item, mgr_a, mgr_b)
        for item in sends_b_items:
            _update_pick_ownership(ownership, item, mgr_b, mgr_a)

        added += 1

    if added > 0:
        trades_data["trades"] = existing_trades
        trades_data["draft_pick_ownership"] = ownership
        with open(trades_path, "w", encoding="utf-8") as f:
            json.dump(trades_data, f, indent=2, ensure_ascii=False)

    return added


def _parse_pick_string(item: str) -> Optional[tuple]:
    """
    Parse a draft pick string like '2026 4th round pick' into (year, round_num).
    Returns None if the string is not a pick.
    """
    # Match patterns like "2026 4th round pick", "2027 1st round pick"
    m = re.match(r"(\d{4})\s+(\d+)(?:st|nd|rd|th)\s+round\s+pick", item, re.IGNORECASE)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return None


def _ordinal(n: int) -> str:
    """Convert integer to ordinal string: 1 -> '1st', 2 -> '2nd', etc."""
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _update_pick_ownership(
    ownership: dict, item: str, sender: str, receiver: str
) -> None:
    """
    Update draft_pick_ownership for a single pick.

    Assumes the pick originally belongs to the sender unless we find
    the sender acquired it from someone else (tracked in ownership).
    """
    parsed = _parse_pick_string(item)
    if not parsed:
        return

    year_str = str(parsed[0])
    round_num = parsed[1]

    # Ensure the year exists in ownership
    if year_str not in ownership:
        ownership[year_str] = {}

    year_picks = ownership[year_str]

    # Figure out the original owner of this pick
    # The sender is sending it, so either:
    # a) It is the sender's own pick (key would be "{round}_{sender}")
    # b) The sender acquired it from someone else (key "{round}_{original}" -> sender)
    # We assume (a) by default -- the sender is trading their own pick

    # Check if the sender actually owns someone else's pick of this round
    # (edge case: sender previously acquired this round from another manager)
    original_owner = sender  # default assumption

    # If the sender's own pick was already traded away, warn
    key_own = f"{round_num}_{sender}"
    if key_own in year_picks and year_picks[key_own] != sender:
        # Sender already traded their own pick away -- they might be
        # sending a pick they acquired from someone else
        for pk, current_owner in year_picks.items():
            if current_owner == sender and pk.startswith(f"{round_num}_"):
                # Found a pick of this round that the sender owns
                original_owner = pk.split("_", 1)[1]
                break

    pick_key = f"{round_num}_{original_owner}"

    if receiver == original_owner:
        # Pick is returning to its original owner -- remove the entry
        year_picks.pop(pick_key, None)
    else:
        year_picks[pick_key] = receiver


def sync_trade_partners_to_records(base_path: Path) -> dict:
    """
    Sync trade partner counts to RECORDS.json by combining:
    1. Historical trades from all_trades.json
    2. Current season trades from TRADES.json
    
    Updates RECORDS.json['all_time']['trade_partners'] with accurate totals.
    
    This ensures the newsletter always cites accurate all-time trade counts
    that include both historical data AND current season trades.
    
    Returns dict of updated trade partner counts, or empty dict if failed.
    """
    from collections import defaultdict
    
    # Find all_trades.json (historical)
    historical_paths = [
        base_path / "data" / "historical" / "all_trades.json",
        base_path / "all_trades.json",
    ]
    historical_path = None
    for p in historical_paths:
        if p.exists():
            historical_path = p
            break
    
    # Find TRADES.json (current season)
    trades_paths = [
        base_path / "config" / "TRADES.json",
        base_path / "TRADES.json",
    ]
    trades_path = None
    for p in trades_paths:
        if p.exists():
            trades_path = p
            break
    
    # Find RECORDS.json
    records_paths = [
        base_path / "config" / "RECORDS.json",
        base_path / "RECORDS.json",
    ]
    records_path = None
    for p in records_paths:
        if p.exists():
            records_path = p
            break
    
    if not records_path:
        return {}
    
    # Count historical trades
    historical_counts = defaultdict(int)
    if historical_path and historical_path.exists():
        try:
            with open(historical_path, encoding="utf-8") as f:
                historical_trades = json.load(f)
            for trade in historical_trades:
                m1 = trade.get("trader_manager", "")
                m2 = trade.get("tradee_manager", "")
                if m1 and m2:
                    # Create canonical key (alphabetical order)
                    pair = "_and_".join(sorted([m1, m2]))
                    historical_counts[pair] += 1
        except Exception as e:
            print(f"Warning: Could not load historical trades: {e}")
    
    # Count current season trades
    current_counts = defaultdict(int)
    if trades_path and trades_path.exists():
        try:
            with open(trades_path, encoding="utf-8") as f:
                trades_data = json.load(f)
            for trade in trades_data.get("trades", []):
                m1 = trade.get("side_a", {}).get("manager", "")
                m2 = trade.get("side_b", {}).get("manager", "")
                if m1 and m2:
                    pair = "_and_".join(sorted([m1, m2]))
                    current_counts[pair] += 1
        except Exception as e:
            print(f"Warning: Could not load current season trades: {e}")
    
    # Combine counts
    all_pairs = set(historical_counts.keys()) | set(current_counts.keys())
    combined = {}
    for pair in all_pairs:
        combined[pair] = historical_counts.get(pair, 0) + current_counts.get(pair, 0)
    
    # Update RECORDS.json
    try:
        with open(records_path, encoding="utf-8") as f:
            records = json.load(f)
        
        if "all_time" not in records:
            records["all_time"] = {}
        
        records["all_time"]["trade_partners"] = combined
        
        with open(records_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2, ensure_ascii=False)
        
        return combined
    except Exception as e:
        print(f"Warning: Could not update RECORDS.json: {e}")
        return {}


def load_trades_context(base_path: Path, current_week: int) -> Optional[dict]:
    """
    Load TRADES.json from config/ and build a context block for the stats report.
    
    Returns a dict with:
    - "season_trades": list of all trades this season (for narrative context)
    - "this_week_trades": list of trades that happened during current_week
    - "recent_trades": list of trades from the last 2 weeks (for "recent" references)
    - "draft_pick_ownership": formatted ownership map showing non-default picks
    - "draft_pick_summary": per-manager summary of pick assets
    
    Returns None if TRADES.json is not found.
    """
    # Search for TRADES.json in common locations
    candidates = [
        base_path / "config" / "TRADES.json",
        base_path / "TRADES.json",
    ]
    
    trades_path = None
    for c in candidates:
        if c.exists():
            trades_path = c
            break
    
    if not trades_path:
        return None
    
    try:
        with open(trades_path, encoding="utf-8") as f:
            trades_data = json.load(f)
    except Exception:
        return None
    
    all_trades = trades_data.get("trades", [])
    ownership = trades_data.get("draft_pick_ownership", {})
    
    # Categorize trades by timing
    this_week = [t for t in all_trades if t.get("week") == current_week]
    recent = [t for t in all_trades if current_week - 2 <= t.get("week", 0) <= current_week]
    
    # Build per-manager pick summary
    managers = MANAGERS
    years = sorted(ownership.keys())
    # Filter out metadata keys
    years = [y for y in years if y.startswith("20")]
    
    pick_summary = {m: {"owns": [], "traded_away": []} for m in managers}
    
    for year in years:
        year_picks = ownership.get(year, {})
        for pick_key, current_owner in year_picks.items():
            # pick_key format: "Round_OriginalOwner" e.g. "1_Garrett"
            parts = pick_key.split("_", 1)
            if len(parts) != 2:
                continue
            round_num, original_owner = parts
            
            # This pick changed hands
            pick_label = f"{year} Round {round_num} ({original_owner}'s)"
            pick_summary[current_owner]["owns"].append(pick_label)
            pick_summary[original_owner]["traded_away"].append(
                f"{year} Round {round_num} -> {current_owner}"
            )
    
    return {
        "season_trades": all_trades,
        "this_week_trades": this_week,
        "recent_trades": recent,
        "draft_pick_ownership": ownership,
        "draft_pick_summary": pick_summary,
    }


def get_key_players_for_matchup(
    manager: str,
    week: int,
    rosters: dict,
    player_projs: dict,
    injured_players: list,
) -> list[dict]:
    """
    Pick 2 key players for a manager's team in a betting line preview.
    
    - Player 1: Best available player by proj FPPG (excluding out/season-long injuries)
    - Player 2: Rotates through #2-#5 by proj FPPG based on week number
    - Returning players are eligible and annotated with return info
    
    Args:
        manager: Manager name
        week: Preview week number (used for rotation)
        rosters: Dict mapping manager -> list of player names
        player_projs: Dict mapping player name -> projected FPPG
        injured_players: List of injured player dicts from current_team_health
    
    Returns:
        List of 2 dicts: [{"name": str, "proj": float, "returning": bool, "return_note": str}, ...]
    """
    roster = rosters.get(manager, [])
    if isinstance(roster, str):
        roster = []
    
    # Build injury lookup: {player_name: {status, return_note}}
    injury_lookup = {}
    for ip in injured_players:
        player = ip.get("player", "")
        status = ip.get("status", "out")
        notes = ip.get("notes", "")
        return_games = ip.get("return_games")
        total_games = ip.get("total_week_games")
        
        return_note = ""
        if status == "returning" and return_games and total_games:
            return_note = f"returning {return_games}/{total_games} games"
        
        injury_lookup[player] = {
            "status": status,
            "return_note": return_note,
        }
    
    # Filter and rank eligible players
    eligible = []
    for player in roster:
        proj = player_projs.get(player)
        if proj is None:
            continue
        
        inj = injury_lookup.get(player)
        if inj:
            if inj["status"] in ("out", "season-long"):
                continue  # Disqualified
            # "returning" players are eligible
            eligible.append({
                "name": player,
                "proj": proj,
                "returning": True,
                "return_note": inj["return_note"],
            })
        else:
            eligible.append({
                "name": player,
                "proj": proj,
                "returning": False,
                "return_note": "",
            })
    
    # Sort by proj FPPG descending
    eligible.sort(key=lambda x: x["proj"], reverse=True)
    
    if not eligible:
        return []
    
    # Player 1: best available
    key_players = [eligible[0]]
    
    # Player 2: rotate through #2-#5 based on week number
    if len(eligible) >= 2:
        candidates = eligible[1:5]  # Up to 4 candidates (ranks 2-5)
        rotation_idx = week % len(candidates)
        key_players.append(candidates[rotation_idx])
    
    return key_players


# =============================================================================
# LAST WEEK RECAP PARSING (fallback for upset detection)
# =============================================================================

def parse_last_week_predictions(recap_path: Path) -> dict:
    """
    Parse LAST_WEEK_RECAP.md to extract betting line predictions.
    
    Returns dict mapping matchup tuple to prediction info:
    {
        ("Nick", "Benton"): {"favorite": "Benton", "spread": -16.5, "win_prob": None},
        ("Hayden", "Garrett"): {"favorite": "Hayden", "spread": -54.0, "win_prob": 57.1},
    }
    """
    predictions = {}
    
    if not recap_path.exists():
        return predictions
    
    try:
        content = recap_path.read_text(encoding="utf-8")
    except Exception:
        return predictions
    
    # Look for the Callbacks Planted section
    callbacks_match = re.search(
        r"## Callbacks Planted.*?\n(.*?)(?=\n## |\Z)", 
        content, 
        re.DOTALL | re.IGNORECASE
    )
    
    if not callbacks_match:
        return predictions
    
    callbacks_section = callbacks_match.group(1)
    
    # Parse each line looking for matchup predictions
    # Pattern: **Nick vs Benton:** ... Nick -36.5 spread ... or ... Benton -16.5 ...
    # FIXED: Use an explicit alternation built from MANAGERS so multi-word
    # names ("Mary Jane", "De'Aaron") are matched correctly. \w+ would only
    # capture the first word.
    managers = MANAGERS
    # Sort longest-first so the regex engine matches the longest manager name
    # before a shorter prefix.
    _mgr_alt = "|".join(re.escape(m) for m in sorted(managers, key=len, reverse=True))
    matchup_re = re.compile(
        rf"\*\*({_mgr_alt})\s+vs\.?\s+({_mgr_alt})\*\*",
        re.IGNORECASE,
    )

    for line in callbacks_section.split("\n"):
        # Find matchup in line (e.g., "Nick vs Benton")
        matchup_match = matchup_re.search(line)
        if not matchup_match:
            continue

        # Normalize to canonical casing via MANAGERS lookup
        _mgr_by_lower = {m.lower(): m for m in managers}
        mgr_a = _mgr_by_lower.get(matchup_match.group(1).lower())
        mgr_b = _mgr_by_lower.get(matchup_match.group(2).lower())

        if mgr_a not in managers or mgr_b not in managers:
            continue

        # Find the spread - look for patterns like "Nick -36.5" or "Benton -16.5 spread"
        # Also handle "revised to Benton -16.5"
        # FIXED: Same MANAGERS-based alternation here.
        spread_patterns = [
            rf"revised to ({_mgr_alt})\s+(-?\d+\.?\d*)",   # "revised to Benton -16.5"
            rf"({_mgr_alt})\s+(-\d+\.?\d*)\s*spread",      # "Hayden -54.0 spread"
            rf"({_mgr_alt})\s+(-\d+\.?\d*)\s*\)",          # "Nick -36.5)"
        ]

        favorite = None
        spread = None

        for pattern in spread_patterns:
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                fav_name = _mgr_by_lower.get(match.group(1).lower())
                if fav_name in managers:
                    favorite = fav_name
                    spread = float(match.group(2))
                    break
        
        # If no spread found, check for "coin-flip" or similar
        if not favorite and "coin-flip" in line.lower():
            favorite = None  # True coin flip
            spread = 0
        
        # Look for win probability
        win_prob = None
        prob_match = re.search(r"(\d+\.?\d*)\s*%\s*win\s*prob", line, re.IGNORECASE)
        if prob_match:
            win_prob = float(prob_match.group(1))
        
        if favorite or spread is not None:
            # Store with sorted tuple key for consistent lookup
            key = tuple(sorted([mgr_a, mgr_b]))
            predictions[key] = {
                "favorite": favorite,
                "spread": spread,
                "win_prob": win_prob,
                "mgr_a": mgr_a,
                "mgr_b": mgr_b,
            }
    
    return predictions


# =============================================================================
# DATA LOADING
# =============================================================================

def load_stats_report(path: Path) -> dict:
    """Load the stats report JSON."""
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def build_lookup_tables(data: dict) -> dict:
    """Build lookup tables for cross-referencing data."""
    lookups = {}
    
    # Manager -> standings record
    lookups["standings"] = {
        s["manager"]: s["record"] 
        for s in data["current_standings"]
    }
    
    # Manager -> team stats (games_played, team_fppg, etc.)
    lookups["team_stats"] = {
        ts["manager"]: ts 
        for ts in data["team_stats"]
    }
    
    # Manager -> report card
    lookups["report_cards"] = {
        rc["manager"]: rc 
        for rc in data["report_cards"]
    }
    
    # Manager -> power ranking
    lookups["power_rankings"] = {
        pr["manager"]: pr 
        for pr in (data.get("power_rankings") or [])
    }
    
    # Manager -> season injury burden
    lookups["injury_burden"] = data["season_injury_burden"]
    
    # Manager -> current team health
    lookups["team_health"] = data["current_team_health"]["teams"]
    
    # Manager -> scoring trends
    lookups["scoring_trends"] = data["scoring_trends"]
    
    # Manager -> current streaks
    lookups["streaks"] = data["current_streaks"]
    
    return lookups


# =============================================================================
# SECTION FORMATTERS
# =============================================================================

def format_section_1_matchup_summaries(data: dict, lookups: dict) -> str:
    """Format Section 1: Matchup Summaries."""
    lines = ["## SECTION 1: MATCHUP SUMMARIES\n"]
    
    for i, matchup in enumerate(data["matchup_summaries"], 1):
        mgr_a = matchup["manager_a"]
        mgr_b = matchup["manager_b"]
        
        # Get cross-referenced data
        record_a = lookups["standings"].get(mgr_a, "?-?")
        record_b = lookups["standings"].get(mgr_b, "?-?")
        ts_a = lookups["team_stats"].get(mgr_a, {})
        ts_b = lookups["team_stats"].get(mgr_b, {})
        pr_a = lookups["power_rankings"].get(mgr_a, {})
        pr_b = lookups["power_rankings"].get(mgr_b, {})
        ib_a = lookups["injury_burden"].get(mgr_a, {})
        ib_b = lookups["injury_burden"].get(mgr_b, {})
        th_a = lookups["team_health"].get(mgr_a, {})
        th_b = lookups["team_health"].get(mgr_b, {})
        st_a = lookups["scoring_trends"].get(mgr_a, {})
        st_b = lookups["scoring_trends"].get(mgr_b, {})
        
        stats_a = matchup["stats_a"]
        stats_b = matchup["stats_b"]
        
        lines.append(f"### Matchup {i}: {mgr_a} vs {mgr_b}\n")
        
        # Result
        lines.append(f"- Winner: {matchup['winner']}")
        lines.append(f"- Score: {matchup['score_a']:.2f} -> {matchup['score_b']:.2f}")
        lines.append(f"- Margin: {matchup['margin']:.2f}")
        
        # Series
        season_a = matchup["season_series"].get(mgr_a, 0)
        season_b = matchup["season_series"].get(mgr_b, 0)
        alltime_a = matchup["all_time_series"].get(mgr_a, 0)
        alltime_b = matchup["all_time_series"].get(mgr_b, 0)
        lines.append(f"- Season series: {mgr_a} {season_a}, {mgr_b} {season_b}")
        lines.append(f"- All-time: {mgr_a} {alltime_a}, {mgr_b} {alltime_b}")
        
        # H2H streak
        if matchup.get("h2h_streak_holder"):
            lines.append(f"- H2H streak: {matchup['h2h_streak_holder']} has won {matchup['h2h_streak_length']} straight")
        
        lines.append("")
        
        # Team A stats
        lines.append(f"**Team A: {mgr_a} ({record_a})**")
        lines.append(f"- Total FP: {stats_a['total_fp']:.2f}")
        lines.append(f"- Games played: {ts_a.get('games_played', '?')} | Team FPPG: {ts_a.get('team_fppg', '?')}")
        lines.append(f"- Efficiency: {stats_a['efficiency_pct']:.1f}%")
        # Get positional breakdown from team_stats (includes games and FPPG)
        pos_a = ts_a.get("positional_stats", {})
        if pos_a:
            g = pos_a.get("guard", {})
            f = pos_a.get("forward", {})
            c = pos_a.get("center", {})
            lines.append(f"- Positional: G: {g.get('total_fp', 0):.2f} ({g.get('fppg', 0):.1f}/g, {g.get('games', 0)}g) | F: {f.get('total_fp', 0):.2f} ({f.get('fppg', 0):.1f}/g, {f.get('games', 0)}g) | C: {c.get('total_fp', 0):.2f} ({c.get('fppg', 0):.1f}/g, {c.get('games', 0)}g)")
        else:
            lines.append(f"- Positional: G: {stats_a['guard_fp']:.2f} | F: {stats_a['forward_fp']:.2f} | C: {stats_a['center_fp']:.2f}")
        
        # Best performers
        if stats_a.get("best_performers"):
            lines.append("- Best performers:")
            for p in stats_a["best_performers"][:3]:
                lines.append(f"  - {p['name']}: {p['fp']:.2f} FP ({p['games']} games)")
        
        # Worst performer
        if stats_a.get("worst_performer") and stats_a["worst_performer"].get("name"):
            wp = stats_a["worst_performer"]
            lines.append(f"- Worst performer: {wp['name']} ({wp['fppg']:.2f} FPPG, {wp['games']} games)")
        
        # Injury breakdown
        lines.append(f"- Games lost to injury: {stats_a['games_lost_to_injury']}")
        if stats_a.get("injury_breakdown"):
            breakdown = ", ".join(
                f"{ib['player']} ({ib['games']})" 
                for ib in stats_a["injury_breakdown"]
            )
            lines.append(f"- Injury breakdown: {breakdown}")
        
        lines.append("")
        
        # Team B stats (same structure)
        lines.append(f"**Team B: {mgr_b} ({record_b})**")
        lines.append(f"- Total FP: {stats_b['total_fp']:.2f}")
        lines.append(f"- Games played: {ts_b.get('games_played', '?')} | Team FPPG: {ts_b.get('team_fppg', '?')}")
        lines.append(f"- Efficiency: {stats_b['efficiency_pct']:.1f}%")
        # Get positional breakdown from team_stats (includes games and FPPG)
        pos_b = ts_b.get("positional_stats", {})
        if pos_b:
            g = pos_b.get("guard", {})
            f = pos_b.get("forward", {})
            c = pos_b.get("center", {})
            lines.append(f"- Positional: G: {g.get('total_fp', 0):.2f} ({g.get('fppg', 0):.1f}/g, {g.get('games', 0)}g) | F: {f.get('total_fp', 0):.2f} ({f.get('fppg', 0):.1f}/g, {f.get('games', 0)}g) | C: {c.get('total_fp', 0):.2f} ({c.get('fppg', 0):.1f}/g, {c.get('games', 0)}g)")
        else:
            lines.append(f"- Positional: G: {stats_b['guard_fp']:.2f} | F: {stats_b['forward_fp']:.2f} | C: {stats_b['center_fp']:.2f}")
        
        if stats_b.get("best_performers"):
            lines.append("- Best performers:")
            for p in stats_b["best_performers"][:3]:
                lines.append(f"  - {p['name']}: {p['fp']:.2f} FP ({p['games']} games)")
        
        if stats_b.get("worst_performer") and stats_b["worst_performer"].get("name"):
            wp = stats_b["worst_performer"]
            lines.append(f"- Worst performer: {wp['name']} ({wp['fppg']:.2f} FPPG, {wp['games']} games)")
        
        lines.append(f"- Games lost to injury: {stats_b['games_lost_to_injury']}")
        if stats_b.get("injury_breakdown"):
            breakdown = ", ".join(
                f"{ib['player']} ({ib['games']})" 
                for ib in stats_b["injury_breakdown"]
            )
            lines.append(f"- Injury breakdown: {breakdown}")
        
        lines.append("")
        
        # Championships
        champ_a = pr_a.get("championships", 0)
        champ_b = pr_b.get("championships", 0)
        lines.append(f"**Championships:** {mgr_a} {champ_a}, {mgr_b} {champ_b}")
        lines.append("")
        
        # Season Injury Context
        lines.append("**Season Injury Context:**")
        lines.append(f"- {mgr_a}: {ib_a.get('total_injury_burden_pct', 0):.1f}% total burden | Total injury games: {ib_a.get('total_injury_games', 0)} | Games lost to injury (non-IL): {ib_a.get('non_il_injury_games', 0)} | IL games: {ib_a.get('il_injury_games', 0)}")
        if ib_a.get("il_players"):
            top_il = ib_a["il_players"][0]
            lines.append(f"  - Top IL player: {top_il['player']} ({top_il['games']} games)")
        lines.append(f"  - NOTE: 'Games lost to injury' = non-IL starter slot injuries only. IL games are tracked separately.")
        
        lines.append(f"- {mgr_b}: {ib_b.get('total_injury_burden_pct', 0):.1f}% total burden | Total injury games: {ib_b.get('total_injury_games', 0)} | Games lost to injury (non-IL): {ib_b.get('non_il_injury_games', 0)} | IL games: {ib_b.get('il_injury_games', 0)}")
        if ib_b.get("il_players"):
            top_il = ib_b["il_players"][0]
            lines.append(f"  - Top IL player: {top_il['player']} ({top_il['games']} games)")
        # (NOTE already printed above for mgr_a, no need to repeat)
        lines.append("")
        
        # Current Team Health
        health_week = data["current_team_health"].get("as_of_week", "?")
        lines.append(f"**Current Team Health (entering Week {health_week}):**")
        
        for mgr, th in [(mgr_a, th_a), (mgr_b, th_b)]:
            health_pct = th.get("health_pct", 100)
            injured_fppg = th.get("injured_fppg", 0)
            lines.append(f"- {mgr}: {health_pct:.1f}% healthy ({injured_fppg:.1f} proj FPPG out)")
            
            injured = th.get("injured_players", [])
            out_players = [p for p in injured if p.get("status") == "out" and not p.get("is_season_long")]
            returning = [p for p in injured if p.get("status") == "returning"]
            season_long = [p for p in injured if p.get("is_season_long")]
            
            if out_players or returning or season_long:
                parts = []
                for p in out_players:
                    parts.append(f"{p['player']} ({p['proj_fppg']:.1f} FPPG, {p['remaining_weeks']} weeks remaining, status: out)")
                for p in returning:
                    parts.append(f"{p['player']} ({p['proj_fppg']:.1f} FPPG, returning {p.get('return_games', '?')}/{p.get('total_week_games', '?')} games)")
                for p in season_long:
                    parts.append(f"{p['player']} (season-long)")
                
                lines.append(f"  - Injured: {'; '.join(parts) if parts else 'None'}")
        
        lines.append("")
        
        # Scoring Trends
        lines.append("**Scoring Trends:**")
        for mgr, st in [(mgr_a, st_a), (mgr_b, st_b)]:
            last3 = st.get("last_3_avg", 0)
            season = st.get("season_avg", 0)
            trend = st.get("trend", "stable")
            desc = st.get("trend_description", "")
            traj = st.get("trajectory", "")
            lines.append(f"- {mgr}: Last 3 avg: {last3:.1f} | Season avg: {season:.1f} | Trend: {trend} | Trajectory: {traj}")
        
        lines.append("\n---\n")
    
    return "\n".join(lines)


def format_section_2_report_cards(data: dict, lookups: dict, prev_grades: Optional[dict] = None) -> str:
    """Format Section 2: Report Cards."""
    lines = ["## SECTION 2: REPORT CARDS\n"]
    
    # Sort by grade (B+ > B > B- > C+ > etc.)
    grade_order = {"A+": 0, "A": 1, "A-": 2, "B+": 3, "B": 4, "B-": 5, 
                   "C+": 6, "C": 7, "C-": 8, "D+": 9, "D": 10, "D-": 11, "F": 12}
    
    sorted_cards = sorted(
        data["report_cards"],
        key=lambda x: grade_order.get(x.get("letter_grade", "C"), 7)
    )
    
    for rc in sorted_cards:
        mgr = rc["manager"]
        team = rc["team_name"]
        grade = rc.get("letter_grade", "?")
        prev = prev_grades.get(mgr) if prev_grades else None
        ib = lookups["injury_burden"].get(mgr, {})
        
        # Header with optional previous grade
        if prev and prev != grade:
            lines.append(f"### {mgr} ({team}) -> {grade} (previously {prev})")
        else:
            lines.append(f"### {mgr} ({team}) -> {grade}")
        lines.append(f"- Record: {rc.get('record', '?-?')}")
        lines.append(f"- Weekly FP: {rc['weekly_fp']:.2f}")
        lines.append(f"- Efficiency: {rc['efficiency_pct']:.1f}%")
        lines.append(f"- Games lost to injury: {rc['games_lost_to_injury']}")
        lines.append(f"- Games left on bench: {rc['games_left_on_bench']}")
        blunders = rc.get("blunders", 0)
        blunder_pts = rc.get("blunder_points", 0.0)
        if blunders > 0:
            lines.append(f"- Blunders: {blunders} ({blunder_pts:.1f} FP wasted)")
        else:
            lines.append(f"- Blunders: 0")
        
        # Waiver info
        waiver_count = rc.get("waiver_adds_count", 0)
        waiver_fp = rc.get("waiver_fp_total", 0)
        fp_per_add = rc.get("fp_per_add", 0)
        waiver_games = rc.get("waiver_games", 0)
        waiver_fppg = rc.get("waiver_fppg", 0)
        if waiver_count > 0:
            lines.append(f"- Waiver adds: {waiver_count} adds, {waiver_fp:.2f} total FP over {waiver_games} games ({waiver_fppg:.2f} fppg, {fp_per_add:.1f} FP per add)")
        else:
            lines.append("- Waiver adds: None")
        
        # Scheduled games context + utilization rate
        sched = rc.get("scheduled_games", 0)
        avg = rc.get("scheduled_games_league_avg", 0)
        starter_games = rc.get("total_starter_games", 0)
        util_rate = (starter_games / sched * 100) if sched > 0 else 0
        lines.append(f"- Scheduled games: {sched} (league avg: {avg:.1f})")
        lines.append(f"- Total starter games: {starter_games}")
        lines.append(f"- Utilization rate: {util_rate:.1f}% ({starter_games}/{sched})")
        
        lines.append("")
        
        # Season injury context
        lines.append("**Season Injury Context:**")
        burden = ib.get("total_injury_burden_pct", 0)
        total_games = ib.get("total_injury_games", 0)
        non_il_games = ib.get("non_il_injury_games", 0)
        il_games = ib.get("il_injury_games", 0)
        
        lines.append(f"- {burden:.1f}% total burden | Total injury games: {total_games} | Games lost to injury (non-IL): {non_il_games} | IL games: {il_games}")
        
        # IL breakdown
        if ib.get("il_players"):
            il_parts = ", ".join(
                f"{p['player']} {p['games']}" 
                for p in ib["il_players"][:3]
            )
            lines.append(f"- Top IL players: {il_parts}")
        
        lines.append(f"- NOTE: 'Games lost to injury' = non-IL starter slot injuries only. IL games are tracked separately.")
        
        lines.append("\n---\n")
    
    return "\n".join(lines)


def format_section_3_betting_lines(data: dict, lookups: dict, player_projs: Optional[dict] = None) -> str:
    """Format Section 3: Betting Lines (Looking Ahead)."""
    la = data.get("looking_ahead")
    if not la:
        return "## SECTION 3: BETTING LINES\n\nNo betting data available.\n"
    
    preview_week = la.get("week", "?")
    preview_week_int = int(preview_week) if isinstance(preview_week, (int, float)) else 0
    lines = [f"## SECTION 3: BETTING LINES (Week {preview_week} Preview)\n"]
    
    # Schedule strength data (if available) -- build_schedule_strength() can return None.
    sched_str = data.get("schedule_strength") or {}
    upcoming_sched = sched_str.get("upcoming_week") or {}
    ros_sched = sched_str.get("rest_of_season") or {}
    sched_managers = upcoming_sched.get("managers") or {}
    ros_managers = ros_sched.get("managers") or {}
    
    # League-wide schedule summary (if data exists)
    if sched_managers:
        lines.append("**Schedule Strength This Week:**")
        sorted_by_startable = sorted(
            sched_managers.items(),
            key=lambda x: -x[1].get("startable_games", 0),
        )
        sched_parts = []
        for mgr, sd in sorted_by_startable:
            sched_parts.append(f"{mgr} {sd.get('startable_games', '?')} startable games ({sd.get('healthy_games', '?')} healthy, {sd.get('total_games', '?')} total)")
        lines.append(f"- {' | '.join(sched_parts)}")
        
        if ros_managers:
            ros_sorted = sorted(
                ros_managers.items(),
                key=lambda x: -x[1].get("total_games", 0),
            )
            ros_parts = [f"{mgr} {sd.get('total_games', '?')}" for mgr, sd in ros_sorted]
            lines.append(f"- ROS total games: {' | '.join(ros_parts)}")
        lines.append("")
    
    # Build injury lookup from current_team_health (any chained section can be None).
    cth_teams = (data.get("current_team_health") or {}).get("teams") or {}
    rosters = data.get("rosters") or {}

    for i, preview in enumerate((la.get("matchup_previews") or []), 1):
        mgr_a = preview["manager_a"]
        mgr_b = preview["manager_b"]
        team_a = preview["team_name_a"]
        team_b = preview["team_name_b"]
        bl = preview["betting_line"]
        
        lines.append(f"### Matchup {i}: {mgr_a} vs {mgr_b} (Week {preview_week})")
        lines.append(f"- Teams: {team_a} vs {team_b}")
        lines.append(f"- Spread: {mgr_a} {bl['spread_a']:+.1f} ({mgr_b} {bl['spread_b']:+.1f})")
        lines.append(f"- O/U: {bl['over_under']}")
        lines.append(f"- Win Prob: {mgr_a} {bl['win_prob_a']:.2f}% | {mgr_b} {bl['win_prob_b']:.2f}%")
        lines.append(f"- Moneyline: {mgr_a} {bl['moneyline_a']:+d} | {mgr_b} {bl['moneyline_b']:+d}")
        lines.append(f"- Avg Score: {mgr_a} {bl['avg_score_a']:.2f} | {mgr_b} {bl['avg_score_b']:.2f}")
        
        # Series
        lines.append(f"- Season series: {mgr_a} {preview.get('series_a_wins', 0)}, {mgr_b} {preview.get('series_b_wins', 0)}")
        lines.append(f"- All-time: {mgr_a} {preview.get('all_time_a_wins', 0)}, {mgr_b} {preview.get('all_time_b_wins', 0)}")
        
        # H2H streak
        if preview.get("h2h_streak_holder"):
            lines.append(f"- H2H streak: {preview['h2h_streak_holder']} has won {preview['h2h_streak_length']} straight")
        
        # Notable injuries
        inj_a = preview.get("notable_injuries_a", [])
        inj_b = preview.get("notable_injuries_b", [])
        if inj_a:
            inj_str = "; ".join(f"{p['player']} ({p.get('notes', 'injured')})" for p in inj_a[:3])
            lines.append(f"- Notable injuries {mgr_a}: {inj_str}")
        if inj_b:
            inj_str = "; ".join(f"{p['player']} ({p.get('notes', 'injured')})" for p in inj_b[:3])
            lines.append(f"- Notable injuries {mgr_b}: {inj_str}")
        
        # Implications
        if preview.get("implications"):
            lines.append(f"- Implications: \"{preview['implications']}\"")
        
        # Key players (2 per team, with rotation for player #2)
        if player_projs and rosters:
            injured_a = cth_teams.get(mgr_a, {}).get("injured_players", [])
            injured_b = cth_teams.get(mgr_b, {}).get("injured_players", [])
            
            kps_a = get_key_players_for_matchup(mgr_a, preview_week_int, rosters, player_projs, injured_a)
            kps_b = get_key_players_for_matchup(mgr_b, preview_week_int, rosters, player_projs, injured_b)
            
            if len(kps_a) >= 2 and len(kps_b) >= 2:
                # Format each player with optional return annotation
                def fmt_player(kp):
                    s = f"{kp['name']} ({kp['proj']:.1f})"
                    if kp["returning"] and kp["return_note"]:
                        s += f" [{kp['return_note']}]"
                    return s
                
                side_a = f"{fmt_player(kps_a[0])} and {fmt_player(kps_a[1])}"
                side_b = f"{fmt_player(kps_b[0])} and {fmt_player(kps_b[1])}"
                lines.append(f"- Key matchup: {side_a} vs {side_b}")
            elif len(kps_a) >= 1 and len(kps_b) >= 1:
                # Fallback: just 1 player each
                lines.append(f"- Key matchup: {kps_a[0]['name']} ({kps_a[0]['proj']:.1f} proj FPPG) vs {kps_b[0]['name']} ({kps_b[0]['proj']:.1f} proj FPPG)")
        else:
            # Fallback to JSON's key_player fields if no projections loaded
            kp_a = preview.get("key_player_a")
            kp_b = preview.get("key_player_b")
            if kp_a and kp_b:
                proj_a = preview.get("key_player_a_proj", 0)
                proj_b = preview.get("key_player_b_proj", 0)
                lines.append(f"- Key matchup: {kp_a} ({proj_a:.1f} proj FPPG) vs {kp_b} ({proj_b:.1f} proj FPPG)")
        
        lines.append("")
        
        # Positional matchups
        lines.append("**Positional Matchups:**")
        for pos_key in ["guard_matchup", "forward_matchup", "center_matchup"]:
            pm = preview.get(pos_key)
            if pm:
                pos = pm["position"]
                fp_a = pm["manager_a_fp"]
                fp_b = pm["manager_b_fp"]
                adv = pm["advantage"]
                margin = pm["advantage_margin"]
                players_a = ", ".join(pm.get("manager_a_players", [])[:3])
                players_b = ", ".join(pm.get("manager_b_players", [])[:3])
                lines.append(f"- {pos}: {mgr_a} {fp_a:.2f} vs {mgr_b} {fp_b:.2f} -> Advantage: {adv} (+{margin:.2f})")
                lines.append(f"  - {mgr_a} players: {players_a}")
                lines.append(f"  - {mgr_b} players: {players_b}")
        
        lines.append("")
        
        # Scoring trends
        st_a = lookups["scoring_trends"].get(mgr_a, {})
        st_b = lookups["scoring_trends"].get(mgr_b, {})
        lines.append("**Scoring Trends:**")
        lines.append(f"- {mgr_a}: {st_a.get('trend', 'stable')} -> {st_a.get('trend_description', '')}")
        lines.append(f"- {mgr_b}: {st_b.get('trend', 'stable')} -> {st_b.get('trend_description', '')}")
        
        # Schedule strength for this matchup
        if sched_managers:
            sd_a = sched_managers.get(mgr_a, {})
            sd_b = sched_managers.get(mgr_b, {})
            start_a = sd_a.get("startable_games", 0)
            start_b = sd_b.get("startable_games", 0)
            bench_a = sd_a.get("bench_games", 0)
            bench_b = sd_b.get("bench_games", 0)
            diff = abs(start_a - start_b)
            if diff > 0:
                edge_mgr = mgr_a if start_a > start_b else mgr_b
                lines.append(f"\n**Schedule Context:**")
                lines.append(f"- {mgr_a}: {start_a} startable games ({bench_a} bench overflow) | {mgr_b}: {start_b} startable games ({bench_b} bench overflow)")
                lines.append(f"- {edge_mgr} has a {diff}-game schedule edge this week")
            else:
                lines.append(f"\n**Schedule Context:**")
                lines.append(f"- Even schedule: both managers have {start_a} startable games this week")
        
        # Consistency context for this matchup
        consistency = data.get("consistency_scores", {})
        cs_managers = consistency.get("managers", {})
        cs_a = cs_managers.get(mgr_a, {})
        cs_b = cs_managers.get(mgr_b, {})
        if cs_a and cs_b:
            lines.append(f"\n**Scoring Consistency:**")
            lines.append(
                f"- {mgr_a}: {cs_a.get('rating', '?')} "
                f"(CV {cs_a.get('cv', 0):.1f}%, "
                f"avg {cs_a.get('mean_fppg', 0):.1f} FPPG, "
                f"floor {cs_a.get('floor_fppg', 0):.1f} / "
                f"ceiling {cs_a.get('ceiling_fppg', 0):.1f}) "
                f"[{cs_a.get('recent_trend', 'Stable')}]"
            )
            lines.append(
                f"- {mgr_b}: {cs_b.get('rating', '?')} "
                f"(CV {cs_b.get('cv', 0):.1f}%, "
                f"avg {cs_b.get('mean_fppg', 0):.1f} FPPG, "
                f"floor {cs_b.get('floor_fppg', 0):.1f} / "
                f"ceiling {cs_b.get('ceiling_fppg', 0):.1f}) "
                f"[{cs_b.get('recent_trend', 'Stable')}]"
            )
            # Predictability note
            cv_a = cs_a.get("cv", 0)
            cv_b = cs_b.get("cv", 0)
            if abs(cv_a - cv_b) >= 2.0:
                more_pred = mgr_a if cv_a < cv_b else mgr_b
                less_pred = mgr_b if cv_a < cv_b else mgr_a
                lines.append(
                    f"- {more_pred} is more predictable; "
                    f"{less_pred} has wider boom/bust range"
                )
        
        lines.append("\n---\n")
    
    return "\n".join(lines)


def format_section_4_potw(data: dict, lookups: dict, potw_history: Optional[list] = None, career_stats: Optional[dict] = None) -> str:
    """Format Section 4: Player of the Week."""
    # player_of_week could be None if the section was skipped or errored upstream.
    potw = data.get("player_of_week") or {}
    winner = potw.get("winner") or {}
    mentions = potw.get("honorable_mentions") or []
    
    lines = ["## SECTION 4: PLAYER OF THE WEEK\n"]
    
    if winner:
        lines.append("**Winner:**")
        lines.append(f"- Player: {winner['player_name']}")
        lines.append(f"- Manager: {winner['manager']}")
        lines.append(f"- NBA Team: {winner.get('nba_team', '?')}")
        lines.append(f"- Total FP: {winner['total_fp']:.2f}")
        lines.append(f"- Games: {winner['games']}")
        lines.append(f"- FPPG: {winner['fppg']:.2f}")
        lines.append(f"- vs Projection: {winner.get('efficiency_pct', 0):+.2f}%")
        lines.append(f"- Team contribution: {winner.get('efficiency_contribution', 0):.2f}%")
        
        bg = winner.get("best_game", {})
        if bg:
            lines.append(f"- Best game: {bg.get('fantasy_points', 0):.1f} FP on {bg.get('date', '?')} vs {bg.get('opponent', '?')}")
        
        lines.append("")
    
    if mentions:
        lines.append("**Honorable Mentions:**")
        for i, m in enumerate(mentions, 1):
            bg = m.get("best_game", {})
            best_str = f"Best game: {bg.get('fantasy_points', 0):.1f} FP on {bg.get('date', '?')} vs {bg.get('opponent', '?')}" if bg else ""
            lines.append(f"{i}. {m['player_name']} ({m['manager']}, {m.get('nba_team', '?')}): {m['total_fp']:.2f} FP, {m['games']} games, {m['fppg']:.2f} FPPG")
            if best_str:
                lines.append(f"   - {best_str}")
        lines.append("")
    
    # POTW history from previous weeks this season
    if potw_history:
        lines.append("**Previous POTW Winners (this season):**")
        for entry in potw_history:
            if entry.get("total_fp", 0) > 0:
                lines.append(f"- Week {entry['week']}: {entry['player']} ({entry['manager']}), {entry['total_fp']:.1f} FP ({entry['games']} games)")
            else:
                lines.append(f"- Week {entry['week']}: {entry['player']} ({entry['manager']})")
        lines.append("")
    
    # Context: season + career POTW stats
    if winner and (potw_history or career_stats):
        current_player = winner["player_name"]
        current_mgr = winner["manager"]
        
        # Season context (from history list)
        prev_wins = [e for e in (potw_history or []) if e["player"] == current_player]
        season_total = len(prev_wins) + 1  # +1 for current week
        if prev_wins:
            prev_weeks = ", ".join(f"Week {e['week']}" for e in prev_wins)
            lines.append(f"**Season POTW context:** {current_player} has now won {season_total} POTW awards this season (previous: {prev_weeks})")
        else:
            lines.append(f"**Season POTW context:** This is {current_player}'s first POTW award this season")
        
        # Career context (from all-time data, only if multiple seasons exist)
        if career_stats:
            by_player = career_stats.get("by_player", {})
            by_manager = career_stats.get("by_manager", {})
            
            # Player career total (previous wins + current = career total)
            player_career = by_player.get(current_player, {})
            career_prev = player_career.get("total", 0)  # doesn't include current week yet
            career_total = career_prev + 1
            num_seasons = len(player_career.get("seasons", {}))
            
            if num_seasons > 1:
                season_breakdown = ", ".join(
                    f"{s}: {c}" for s, c in sorted(player_career.get("seasons", {}).items())
                )
                lines.append(f"**Career POTW:** {current_player} has {career_total} career POTW awards ({season_breakdown}, + 1 this week)")
        
        # Manager POTW leaderboard (season)
        mgr_wins = {}
        for e in (potw_history or []):
            mgr_wins[e["manager"]] = mgr_wins.get(e["manager"], 0) + 1
        mgr_wins[current_mgr] = mgr_wins.get(current_mgr, 0) + 1
        
        mgr_summary = ", ".join(f"{m} {c}" for m, c in sorted(mgr_wins.items(), key=lambda x: -x[1]))
        lines.append(f"**POTW by manager (season):** {mgr_summary}")
        
        # Manager career leaderboard (only if multiple seasons)
        if career_stats and len(career_stats.get("by_manager", {}).values()) > 0:
            by_manager = career_stats.get("by_manager", {})
            # Check if any manager has wins across multiple seasons
            any_multi_season = any(len(m.get("seasons", {})) > 1 for m in by_manager.values())
            if any_multi_season:
                career_mgr = {}
                for m, stats in by_manager.items():
                    career_mgr[m] = stats["total"]
                career_mgr[current_mgr] = career_mgr.get(current_mgr, 0) + 1
                career_summary = ", ".join(f"{m} {c}" for m, c in sorted(career_mgr.items(), key=lambda x: -x[1]))
                lines.append(f"**POTW by manager (career):** {career_summary}")
        
        lines.append("")
    
    return "\n".join(lines)


def format_section_5_fun_facts(data: dict, lookups: dict) -> str:
    """Format Section 5: Fun Facts."""
    lines = ["## SECTION 5: FUN FACTS\n"]
    
    # Fun facts from generator
    facts = data.get("fun_facts", [])
    if facts:
        lines.append("**From fun_facts[]:**")
        for i, f in enumerate(facts, 1):
            lines.append(f"{i}. \"{f['text']}\" (category: {f.get('category', 'general')})")
        lines.append("")
    
    # Current streaks
    streaks = lookups["streaks"]
    lines.append("**Season-Longest Streaks (from current_streaks):**")
    lines.append("| Manager | Current Win | Current Loss | Season-Best Win | Season-Best Loss |")
    lines.append("|---------|-------------|--------------|-----------------|------------------|")
    for mgr, s in streaks.items():
        win = s.get("win_streak", 0)
        loss = s.get("loss_streak", 0)
        best_win = s.get("season_longest_win_streak", 0)
        best_win_wks = s.get("season_longest_win_weeks", "")
        best_loss = s.get("season_longest_loss_streak", 0)
        best_loss_wks = s.get("season_longest_loss_weeks", "")
        lines.append(f"| {mgr} | {win} | {loss} | {best_win} (weeks {best_win_wks}) | {best_loss} (weeks {best_loss_wks}) |")
    lines.append("")
    
    # Record updates
    updates = data.get("record_updates", [])
    if updates:
        lines.append("**Record Updates:**")
        for u in updates:
            lines.append(f"- {u['description']}: {u['value']}")
        lines.append("")
    
    # All-time longest streaks for context -- build_all_time_records() can return None.
    atr = data.get("all_time_records") or {}
    win_streaks = atr.get("longest_win_streaks") or {}
    loss_streaks = atr.get("longest_loss_streaks") or {}
    
    if win_streaks:
        lines.append("**All-Time Longest Streaks (for context):**")
        win_parts = ", ".join(f"{m} {s['length']} ({s['season']})" for m, s in win_streaks.items())
        loss_parts = ", ".join(f"{m} {s['length']} ({s['season']})" for m, s in loss_streaks.items())
        lines.append(f"- Win: {win_parts}")
        lines.append(f"- Loss: {loss_parts}")
        lines.append("")
    
    # All-time record benchmarks
    lines.append("**All-Time Record Benchmarks:**")
    
    highest = atr.get("highest_weekly_score", {})
    if highest:
        lines.append(f"- Highest weekly score: {highest.get('score', '?')} ({highest.get('manager', '?')}, {highest.get('season', '?')} Week {highest.get('week', '?')})")
    
    lowest = atr.get("lowest_weekly_score", {})
    if lowest:
        lines.append(f"- Lowest weekly score: {lowest.get('score', '?')} ({lowest.get('manager', '?')}, {lowest.get('season', '?')} Week {lowest.get('week', '?')})")
    
    blowout = atr.get("biggest_blowout", {})
    if blowout:
        lines.append(f"- Biggest blowout: {blowout.get('margin', '?'):.1f} pts ({blowout.get('winner', '?')} over {blowout.get('loser', '?')}, {blowout.get('season', '?')} Week {blowout.get('week', '?')})")
    
    closest = atr.get("closest_game", {})
    if closest:
        lines.append(f"- Closest game: {closest.get('margin', '?'):.1f} pts ({closest.get('winner', '?')} over {closest.get('loser', '?')}, {closest.get('season', '?')} Week {closest.get('week', '?')})")
    
    lines.append("")
    
    # All-time H2H records matrix
    h2h = atr.get("h2h_records", {})
    if h2h:
        lines.append("**All-Time Head-to-Head Records:**")
        lines.append("| Matchup | Record |")
        lines.append("|---------|--------|")
        for matchup, record in h2h.items():
            # Parse the matchup key (e.g., "Hayden_vs_Nick")
            parts = matchup.split("_vs_")
            if len(parts) == 2:
                m1, m2 = parts[0], parts[1]
                w1 = record.get(m1.lower(), 0)
                w2 = record.get(m2.lower(), 0)
                lines.append(f"| {m1} vs {m2} | {m1} {w1} - {w2} {m2} |")
        lines.append("")
    
    # Luck Index (All-Play Expected Wins)
    luck = data.get("luck_index", {})
    historical = data.get("historical_luck", {})
    if luck and luck.get("managers"):
        lines.append("**Luck Index (All-Play Expected Wins):**")
        lines.append("_Expected wins = fraction of league outscored each week, summed across the season._")
        lines.append("_Positive luck = more wins than a balanced schedule would give. Negative = fewer (tougher matchup draws)._")
        lines.append("")
        lines.append("| Manager | Actual | Expected | Luck | Rating | Avg PF | Avg PA | Margin |")
        lines.append("|---------|--------|----------|------|--------|--------|--------|--------|")
        
        # Sort by luck_index descending (luckiest first)
        sorted_mgrs = sorted(
            luck["managers"].values(),
            key=lambda x: x.get("luck_index", 0),
            reverse=True,
        )
        for m in sorted_mgrs:
            lines.append(
                f"| {m['manager']} "
                f"| {m['actual_record']} "
                f"| {m['expected_record']} "
                f"| {m['luck_index']:+.1f} "
                f"| {m['luck_rating']} "
                f"| {m['avg_points_for']:.1f} "
                f"| {m['avg_points_against']:.1f} "
                f"| {m['scoring_margin']:+.1f} |"
            )
        lines.append("")
        lines.append(f"Luckiest: {luck['luckiest']} | Unluckiest: {luck['unluckiest']}")
        lines.append("")
        
        # Historical context if available
        if historical and historical.get("managers"):
            lines.append("**Historical Luck Context:**")
            lines.append("")
            
            # Career totals
            lines.append("_Career Luck (all-time):_")
            lines.append("| Manager | Record | Win% | Expected | Career Luck |")
            lines.append("|---------|--------|------|----------|-------------|")
            sorted_career = sorted(
                historical["managers"].values(),
                key=lambda x: x.get("career_luck", 0),
                reverse=True,
            )
            for m in sorted_career:
                lines.append(
                    f"| {m['manager']} "
                    f"| {m['actual_record']} "
                    f"| {m['win_pct']:.1f}% "
                    f"| {m['expected_wins']:.0f}-{m['total_games']-m['expected_wins']:.0f} "
                    f"| {m['career_luck']:+.1f} |"
                )
            lines.append("")
            
            # Current season rank for each manager
            lines.append("_This season vs career history:_")
            for mgr_name, mgr_data in historical["managers"].items():
                seasons = mgr_data.get("seasons", [])
                current = next((s for s in seasons if s["season"] == CURRENT_SEASON), None)
                if current and seasons:
                    sorted_seasons = sorted(seasons, key=lambda s: -s["luck_index"])
                    rank = next((i+1 for i, s in enumerate(sorted_seasons) if s["season"] == CURRENT_SEASON), 0)
                    luckiest = mgr_data.get("luckiest_season", {})
                    unluckiest = mgr_data.get("unluckiest_season", {})
                    lines.append(
                        f"- {mgr_name}'s {current['luck_index']:+.1f} luck ranks #{rank} of {len(seasons)} career seasons "
                        f"(best: {luckiest.get('season', '?')} {luckiest.get('luck_index', 0):+.1f}, "
                        f"worst: {unluckiest.get('season', '?')} {unluckiest.get('luck_index', 0):+.1f})"
                    )
            lines.append("")
            
            # League records
            luckiest_ever = historical.get("luckiest_single_season")
            unluckiest_ever = historical.get("unluckiest_single_season")
            if luckiest_ever:
                lines.append(f"League record luckiest season: {luckiest_ever['manager']} {luckiest_ever['season']} ({luckiest_ever['luck_index']:+.1f})")
            if unluckiest_ever:
                lines.append(f"League record unluckiest season: {unluckiest_ever['manager']} {unluckiest_ever['season']} ({unluckiest_ever['luck_index']:+.1f})")
            lines.append("")
    
    return "\n".join(lines)


def format_section_6_what_if(data: dict, lookups: dict) -> str:
    """Format Section 6: What If."""
    lines = ["## SECTION 6: WHAT IF\n"]
    
    what_if = data.get("what_if", {})
    manager_analysis = what_if.get("manager_analysis", {})
    notable = what_if.get("notable_swaps", [])
    
    # Notable swaps
    if notable:
        lines.append("**From what_if.notable_swaps[]:**")
        for swap in notable:
            lines.append(f"- {swap}")
        lines.append("")
    else:
        lines.append("**From what_if.notable_swaps[]:**")
        lines.append("- NONE")
        lines.append("")
    
    # Manager analysis
    lines.append("**Manager Analysis:**")
    any_regrets = False
    any_blunders = False
    for mgr, analysis in manager_analysis.items():
        bench = analysis.get("total_bench_points", 0)
        gain = analysis.get("total_potential_gain", 0)
        flip = "YES" if analysis.get("would_flip_matchup") else "NO"
        blunders = analysis.get("blunders", 0)
        blunder_pts = analysis.get("blunder_points", 0.0)
        lines.append(f"- {mgr}: {bench:.1f} bench pts, {gain:.1f} potential gain, would {'flip' if flip == 'YES' else 'NOT flip'}, {blunders} blunder(s) ({blunder_pts:.1f} FP)")
        if bench > 0 or gain > 0:
            any_regrets = True
        if blunders > 0:
            any_blunders = True
            # Detail each blunder
            for b in analysis.get("blunder_details", []):
                lines.append(f"  - BLUNDER: {b['bench_player']} ({b['bench_player_fp']:.1f} FP) left on bench -> {b['available_slot']} slot ({b['dnp_starter']} didn't play) on {b['date']}")
    
    lines.append("")
    
    if not any_regrets and not any_blunders:
        lines.append("**Summary:** All managers left 0 games on the bench. Perfect lineup management -> no regrets this week.")
    
    lines.append("")
    return "\n".join(lines)


def _format_section_7_playoff(data: dict, lookups: dict, playoff_odds: dict, prev_title_odds: Optional[dict] = None) -> str:
    """Format Section 7 for playoff weeks: Championship Odds & Bracket."""
    lines = ["## SECTION 7: PLAYOFF CHAMPIONSHIP ODDS\n"]

    # power_rankings is None in --fast mode; `or []` guards both missing-key and null cases.
    rankings = data.get("power_rankings") or []
    seeds = playoff_odds.get("seeds", {})
    playoff_round = playoff_odds.get("playoff_round", "pre_semis")
    
    # Playoff bracket header
    lines.append(f"**Playoff Round:** {'Semifinals' if playoff_round == 'pre_semis' else 'Finals'}")
    lines.append("")
    
    # Semifinal matchups with win probabilities
    semi_matchups = playoff_odds.get("semi_matchups", [])
    if semi_matchups:
        lines.append("**Semifinal Matchups (Week 22):**")
        lines.append("| Matchup | Seed | Manager | Win Prob | vs | Manager | Seed | Win Prob |")
        lines.append("|---------|------|---------|---------|-----|---------|------|---------|")
        for i, semi in enumerate(semi_matchups, 1):
            lines.append(
                f"| Game {i} | #{semi['seed_a']} | {semi['manager_a']} | {semi['win_prob_a']:.1f}% "
                f"| vs | {semi['manager_b']} | #{semi['seed_b']} | {semi['win_prob_b']:.1f}% |"
            )
        lines.append("")
    
    # Championship probability
    lines.append("**Championship Probability:**")
    lines.append("| Seed | Manager | Team | Record | Champ % | Runner-Up % | 3rd % | 4th % |")
    lines.append("|------|---------|------|--------|---------|-------------|-------|-------|")
    
    fd = playoff_odds.get("finish_distribution", {})
    for pr in rankings:
        mgr = pr["manager"]
        seed = seeds.get(mgr, "?")
        dist = fd.get(mgr, {})
        lines.append(
            f"| #{seed} | {pr['manager']} | {pr['team_name']} | {pr['record']} "
            f"| {dist.get(1, dist.get('1', 0)):.1f}% "
            f"| {dist.get(2, dist.get('2', 0)):.1f}% "
            f"| {dist.get(3, dist.get('3', 0)):.1f}% "
            f"| {dist.get(4, dist.get('4', 0)):.1f}% |"
        )
    lines.append("")
    
    # Most likely championship game
    champ_probs = playoff_odds.get("championship_matchup_probs", {})
    if champ_probs:
        lines.append("**Most Likely Championship Matchup (Week 23):**")
        sorted_matchups = sorted(champ_probs.items(), key=lambda x: x[1], reverse=True)
        for matchup_key, prob in sorted_matchups:
            m1, m2 = matchup_key.split("_vs_")
            lines.append(f"- {m1} vs {m2}: {prob:.1f}%")
        lines.append("")
    
    # Title odds movement from last week
    if prev_title_odds and rankings:
        lines.append("**Championship Odds Movement (vs Last Week):**")
        for pr in rankings:
            mgr = pr["manager"]
            current = pr["title_odds"]
            prev = prev_title_odds.get(mgr)
            if prev is not None:
                delta = current - prev
                if abs(delta) < 0.05:
                    lines.append(f"- {mgr}: {prev:.1f}% -> {current:.1f}% (unchanged)")
                else:
                    lines.append(f"- {mgr}: {prev:.1f}% -> {current:.1f}% ({delta:+.1f}%)")
            else:
                lines.append(f"- {mgr}: {current:.1f}% (new -- entering playoffs)")
        lines.append("")
    
    # Keeper quality (still relevant in playoffs)
    has_keeper_quality = any(pr.get("keeper_quality") is not None for pr in rankings)
    if has_keeper_quality:
        lines.append("**Keeper Quality Scores:**")
        lines.append("| Rank | Manager | Keeper Quality | Top 5 Keepers | Career Record | Championships |")
        lines.append("|------|---------|----------------|---------------|---------------|---------------|")
        for pr in rankings:
            kq = pr.get("keeper_quality")
            kq_str = f"{kq:.1f}" if kq is not None else "-"
            top5 = ", ".join(pr.get("top_5_keepers", [])[:5])
            lines.append(f"| {pr['rank']} | {pr['manager']} | {kq_str} | {top5} | {pr['career_record']} | {pr['championships']} |")
        lines.append("")
    
    # Season injury burden (same as regular season)
    lines.append("**Season Injury Burden:**")
    lines.append("| Manager | Total Burden % | Total Injury Games | Games Lost to Injury (non-IL) | IL Games | Top IL Player |")
    lines.append("|---------|---------------|-------------------|-------------------------------|----------|---------------|")
    
    sorted_ib = sorted(
        lookups["injury_burden"].items(),
        key=lambda x: x[1].get("total_injury_burden_pct", 0),
        reverse=True
    )
    for mgr, ib in sorted_ib:
        total_pct = ib.get("total_injury_burden_pct", 0)
        total_games = ib.get("total_injury_games", 0)
        non_il = ib.get("non_il_injury_games", 0)
        il_games = ib.get("il_injury_games", 0)
        top_il = ib["il_players"][0] if ib.get("il_players") else {"player": "None", "games": 0}
        lines.append(f"| {mgr} | {total_pct:.1f}% | {total_games} | {non_il} | {il_games} | {top_il['player']} ({top_il['games']}) |")
    lines.append("*(Note: Total Injury Games = Games Lost to Injury + IL Games. 'Games Lost to Injury' counts non-IL starter slot injuries. IL Games are tracked separately.)*")
    lines.append("")
    
    # Current team health (critical for playoff context)
    if data.get("current_team_health"):
        health_week = data["current_team_health"].get("as_of_week", "?")
        lines.append(f"**Current Team Health (entering Week {health_week}):**")
        lines.append("| Manager | Health % | Injured FPPG | Key Injured Players |")
        lines.append("|---------|----------|-------------|---------------------|")
        
        sorted_health = sorted(
            lookups.get("team_health", {}).items(),
            key=lambda x: x[1].get("health_pct", 0),
            reverse=True
        )
        for mgr, th in sorted_health:
            health = th.get("health_pct", 100)
            inj_fppg = th.get("injured_fppg", 0)
            injured = th.get("injured_players", [])
            
            key_injured = []
            for p in injured:
                if p.get("is_season_long"):
                    key_injured.append(f"{p['player']} (season)")
                elif p.get("status") == "returning":
                    key_injured.append(f"{p['player']} (returning {p.get('return_games', '?')}/{p.get('total_week_games', '?')})")
                else:
                    key_injured.append(f"{p['player']} ({p['proj_fppg']:.1f}, {p['remaining_weeks']} wks)")
            
            lines.append(f"| {mgr} | {health:.1f}% | {inj_fppg:.1f} | {', '.join(key_injured[:3]) if key_injured else 'None'} |")
        lines.append("")
    
    # Scoring trends (who's hot/cold entering playoffs)
    if lookups.get("scoring_trends"):
        lines.append("**Scoring Trends:**")
        lines.append("| Manager | Last 3 Avg | Season Avg | Trend | Description |")
        lines.append("|---------|-----------|------------|-------|-------------|")
        for mgr, st in lookups["scoring_trends"].items():
            lines.append(f"| {mgr} | {st['last_3_avg']:.1f} | {st['season_avg']:.1f} | {st['trend']} | {st['trend_description']} |")
        lines.append("")
    
    # Returning players (especially important for playoff matchups)
    returning = (data.get("current_team_health") or {}).get("returning_players") or []
    if returning:
        lines.append("**Returning Players:**")
        for r in returning:
            notes = r.get("return_notes", r.get("notes", ""))
            lines.append(f"- {r['player']} ({r['manager']}): {r.get('return_games', '?')}/{r.get('total_week_games', '?')} games -> {notes}")
        lines.append("")
    
    return "\n".join(lines)


def format_section_7_power_rankings(data: dict, lookups: dict, prev_title_odds: Optional[dict] = None) -> str:
    """Format Section 7: Power Rankings (or Playoff Championship Odds)."""
    is_playoff = data.get("is_playoff_week", False)
    playoff_odds = data.get("playoff_odds")
    
    if is_playoff and playoff_odds:
        return _format_section_7_playoff(data, lookups, playoff_odds, prev_title_odds)
    
    lines = ["## SECTION 7: POWER RANKINGS\n"]

    # power_rankings is None in --fast mode; `or []` guards both missing-key and null cases.
    rankings = data.get("power_rankings") or []

    # Current standings
    standings = data.get("current_standings") or []
    if standings:
        lines.append("**Current Standings:**")
        lines.append("| Place | Manager | Record |")
        lines.append("|-------|---------|--------|")
        for i, s in enumerate(standings, 1):
            lines.append(f"| {i} | {s['manager']} | {s['record']} |")
        lines.append("")
    
    # Projected finish distribution
    lines.append("**Projected Finish Distribution:**")
    lines.append("| Team | 1st | 2nd | 3rd | 4th |")
    lines.append("|------|-----|-----|-----|-----|")
    for pr in rankings:
        fd = pr.get("finish_distribution", {})
        lines.append(f"| {pr['team_name']} | {fd.get('1', 0):.2f}% | {fd.get('2', 0):.2f}% | {fd.get('3', 0):.2f}% | {fd.get('4', 0):.2f}% |")
    lines.append("")
    
    # Power rankings table
    has_keeper_quality = any(pr.get("keeper_quality") is not None for pr in rankings)
    lines.append("**Power Rankings:**")
    if has_keeper_quality:
        lines.append("| Rank | Manager | Team | Record | Title Odds | Trend | Keeper Quality | Top 5 Keepers | Expected Record | Career Record | Career Win % | Championships |")
        lines.append("|------|---------|------|--------|------------|-------|----------------|---------------|-----------------|---------------|--------------|---------------|")
        for pr in rankings:
            kq = pr.get("keeper_quality")
            kq_str = f"{kq:.1f}" if kq is not None else "-"
            top5 = ", ".join(pr.get("top_5_keepers", [])[:5])
            lines.append(f"| {pr['rank']} | {pr['manager']} | {pr['team_name']} | {pr['record']} | {pr['title_odds']:.1f}% | {pr['trend']} | {kq_str} | {top5} | {pr['expected_record']} | {pr['career_record']} | {pr['career_win_pct']:.1f}% | {pr['championships']} |")
    else:
        lines.append("| Rank | Manager | Team | Record | Title Odds | Trend | Expected Record | Career Record | Career Win % | Championships |")
        lines.append("|------|---------|------|--------|------------|-------|-----------------|---------------|--------------|---------------|")
        for pr in rankings:
            lines.append(f"| {pr['rank']} | {pr['manager']} | {pr['team_name']} | {pr['record']} | {pr['title_odds']:.1f}% | {pr['trend']} | {pr['expected_record']} | {pr['career_record']} | {pr['career_win_pct']:.1f}% | {pr['championships']} |")
    lines.append("")
    
    # Title odds movement (requires previous week data)
    if prev_title_odds and rankings:
        lines.append("**Title Odds Movement:**")
        for pr in rankings:
            mgr = pr["manager"]
            current = pr["title_odds"]
            prev = prev_title_odds.get(mgr)
            if prev is not None:
                delta = current - prev
                if abs(delta) < 0.05:
                    lines.append(f"- {mgr}: {prev:.1f}% -> {current:.1f}% (unchanged)")
                else:
                    arrow = " -> " if delta > 0 else " -> "
                    lines.append(f"- {mgr}: {prev:.1f}% -> {current:.1f}% ({delta:+.1f}%) {arrow}")
            else:
                lines.append(f"- {mgr}: {current:.1f}% (no previous data)")
        lines.append("")
    
    # Magic number -- find the first manager with a magic_number set
    for _mgr_name in MANAGERS:
        _mgr_pr = lookups["power_rankings"].get(_mgr_name, {})
        if _mgr_pr.get("magic_number"):
            lines.append(f"**Magic Number:** {_mgr_name}'s magic number is {_mgr_pr['magic_number']}")
            break
        lines.append("")
    
    # Season injury burden
    lines.append("**Season Injury Burden:**")
    lines.append("| Manager | Total Burden % | Total Injury Games | Games Lost to Injury (non-IL) | IL Games | Top IL Player |")
    lines.append("|---------|---------------|-------------------|-------------------------------|----------|---------------|")
    
    sorted_ib = sorted(
        lookups["injury_burden"].items(),
        key=lambda x: x[1].get("total_injury_burden_pct", 0),
        reverse=True
    )
    for mgr, ib in sorted_ib:
        top_il = ib["il_players"][0] if ib.get("il_players") else {"player": "None", "games": 0}
        lines.append(f"| {mgr} | {ib.get('total_injury_burden_pct', 0):.1f}% | {ib.get('total_injury_games', 0)} | {ib.get('non_il_injury_games', 0)} | {ib.get('il_injury_games', 0)} | {top_il['player']} ({top_il['games']}) |")
    lines.append("*(Note: Total Injury Games = Games Lost to Injury + IL Games. 'Games Lost to Injury' counts non-IL starter slot injuries. IL Games are tracked separately.)*")
    lines.append("")
    
    # Current team health
    health_week = data["current_team_health"].get("as_of_week", "?")
    lines.append(f"**Current Team Health (entering Week {health_week}):**")
    lines.append("| Manager | Health % | Injured FPPG | Key Injured Players |")
    lines.append("|---------|----------|-------------|---------------------|")
    
    sorted_health = sorted(
        lookups["team_health"].items(),
        key=lambda x: x[1].get("health_pct", 0),
        reverse=True
    )
    for mgr, th in sorted_health:
        health = th.get("health_pct", 100)
        inj_fppg = th.get("injured_fppg", 0)
        injured = th.get("injured_players", [])
        
        key_injured = []
        for p in injured:
            if p.get("is_season_long"):
                key_injured.append(f"{p['player']} (season)")
            elif p.get("status") == "returning":
                key_injured.append(f"{p['player']} (returning {p.get('return_games', '?')}/{p.get('total_week_games', '?')})")
            else:
                key_injured.append(f"{p['player']} ({p['proj_fppg']:.1f}, {p['remaining_weeks']} wks)")
        
        lines.append(f"| {mgr} | {health:.1f}% | {inj_fppg:.1f} | {', '.join(key_injured[:3]) if key_injured else 'None'} |")
    lines.append("")
    
    # Scoring trends
    lines.append("**Scoring Trends:**")
    lines.append("| Manager | Last 3 Avg | Season Avg | Trend | Description |")
    lines.append("|---------|-----------|------------|-------|-------------|")
    for mgr, st in lookups["scoring_trends"].items():
        lines.append(f"| {mgr} | {st['last_3_avg']:.1f} | {st['season_avg']:.1f} | {st['trend']} | {st['trend_description']} |")
    lines.append("")
    
    # Returning players
    returning = data["current_team_health"].get("returning_players", [])
    if returning:
        lines.append("**Returning Players:**")
        for r in returning:
            notes = r.get("return_notes", r.get("notes", ""))
            lines.append(f"- {r['player']} ({r['manager']}): {r.get('return_games', '?')}/{r.get('total_week_games', '?')} games -> {notes}")
        lines.append("")
    
    # Season FPPG stats (for context)
    season_fppg = data.get("season_fppg_stats", {})
    if season_fppg:
        lines.append("**Season FPPG by Position:**")
        lines.append("| Manager | Overall FPPG | G FPPG | F FPPG | C FPPG |")
        lines.append("|---------|-------------|--------|--------|--------|")
        for mgr in MANAGERS:
            stats = season_fppg.get(mgr, {})
            overall = stats.get("fppg", stats.get("season_fppg", 0))
            pos = stats.get("positional", {})
            g_fppg = pos.get("guard", {}).get("fppg", 0)
            f_fppg = pos.get("forward", {}).get("fppg", 0)
            c_fppg = pos.get("center", {}).get("fppg", 0)
            lines.append(f"| {mgr} | {overall:.2f} | {g_fppg:.2f} | {f_fppg:.2f} | {c_fppg:.2f} |")
        lines.append("")
    
    # Consistency scores
    consistency = data.get("consistency_scores", {})
    cs_managers = consistency.get("managers", {})
    if cs_managers:
        lines.append("**Scoring Consistency (Weekly FPPG Volatility):**")
        lines.append("| Manager | CV | Rating | Mean FPPG | Floor | Ceiling | Boom Wks | Bust Wks | Recent Trend |")
        lines.append("|---------|-----|--------|-----------|-------|---------|----------|----------|-------------|")
        sorted_cs = sorted(cs_managers.items(), key=lambda x: x[1].get("cv", 0))
        for mgr, cs in sorted_cs:
            lines.append(
                f"| {mgr} "
                f"| {cs.get('cv', 0):.1f}% "
                f"| {cs.get('rating', '?')} "
                f"| {cs.get('mean_fppg', 0):.2f} "
                f"| {cs.get('floor_fppg', 0):.2f} (Wk {cs.get('floor_week', '?')}) "
                f"| {cs.get('ceiling_fppg', 0):.2f} (Wk {cs.get('ceiling_week', '?')}) "
                f"| {cs.get('boom_weeks', 0)} "
                f"| {cs.get('bust_weeks', 0)} "
                f"| {cs.get('recent_trend', 'Stable')} |"
            )
        lines.append("*(CV = Coefficient of Variation on weekly FPPG. Lower = more consistent. Boom/Bust = weeks > 1 SD from mean.)*")
        lines.append("")
        
        # Player consistency highlights
        has_player_data = any(
            cs.get("most_consistent_player") for cs in cs_managers.values()
        )
        if has_player_data:
            lines.append("**Player Consistency Highlights:**")
            for mgr in MANAGERS:
                cs = cs_managers.get(mgr, {})
                mc_p = cs.get("most_consistent_player")
                mv_p = cs.get("most_volatile_player")
                if mc_p and mv_p:
                    lines.append(
                        f"- {mgr}: Most consistent = {mc_p['player_name']} "
                        f"(CV {mc_p['cv']:.1f}%, {mc_p['mean_fp']:.1f} avg) | "
                        f"Most volatile = {mv_p['player_name']} "
                        f"(CV {mv_p['cv']:.1f}%, {mv_p['mean_fp']:.1f} avg)"
                    )
            lines.append("")
    
    return "\n".join(lines)


def format_section_8_stats_corner(data: dict, lookups: dict) -> str:
    """Format Section 8: Stats Corner."""
    lines = ["## SECTION 8: STATS CORNER\n"]
    
    bw = data.get("best_worst", {})
    # Use `or {}` (not just a default) so a present-but-None section -- e.g.
    # when the live season-performers pull is unavailable -- degrades to empty
    # tables instead of crashing the entire newsletter.
    sp = data.get("season_performers") or {}
    
    # Best games
    best = bw.get("best_games", [])[:5]
    if best:
        lines.append("**Top Performances (best_games):**")
        lines.append("| Rank | Player | NBA Team | Manager | FP | Date | Opponent |")
        lines.append("|------|--------|----------|---------|-----|------|----------|")
        for i, g in enumerate(best, 1):
            lines.append(f"| {i} | {g['player_name']} | {g['nba_team']} | {g['manager']} | {g['fantasy_points']:.2f} | {g['date']} | {g['nba_opponent']} |")
        lines.append("")
    
    # Worst games
    worst = bw.get("worst_games", [])[:5]
    if worst:
        lines.append("**Worst Performances (worst_games):**")
        lines.append("| Rank | Player | NBA Team | Manager | FP | Date | Opponent |")
        lines.append("|------|--------|----------|---------|-----|------|----------|")
        for i, g in enumerate(worst, 1):
            lines.append(f"| {i} | {g['player_name']} | {g['nba_team']} | {g['manager']} | {g['fantasy_points']:.2f} | {g['date']} | {g['nba_opponent']} |")
        lines.append("")
    
    # Season performers
    # Cross-reference with consistency scores for CV and IQR
    consistency = data.get("consistency_scores", {})
    player_cv_lookup = consistency.get("player_lookup", {})

    # Positional Scoring Breakdown -- REMOVED (rendered by Stats Corner viz)

    def format_season_table(title: str, players: list, sort_col: str):
        if not players:
            return []
        result = [f"**{title}:**"]
        # Column order depends on sort_col: FPPG tables swap FPPG and Total FP
        if sort_col == "fppg":
            result.append("| Rank | Player | Fantasy Team | NBA Team | FPPG | Total FP | GP% | MPG | Eff% | CV | IQR (25-75) | Proj FP (ROS) | Proj FPPG (ROS) |")
            result.append("|------|--------|--------------|----------|------|----------|-----|-----|------|-----|-------------|---------------|-----------------|")
        else:
            result.append("| Rank | Player | Fantasy Team | NBA Team | Total FP | FPPG | GP% | MPG | Eff% | CV | IQR (25-75) | Proj FP (ROS) | Proj FPPG (ROS) |")
            result.append("|------|--------|--------------|----------|----------|------|-----|-----|------|-----|-------------|---------------|-----------------|")
        for i, p in enumerate(players[:30], 1):
            pname = p["player_name"]
            pc = player_cv_lookup.get(pname, {})
            cv_str = f"{pc['cv']:.1f}%" if pc else "-"
            iqr_str = f"{pc['q25']:.1f}-{pc['q75']:.1f}" if pc else "-"
            mpg_str = f"{p['mpg']:.1f}" if p.get('mpg') is not None else "-"
            proj_fp_str = f"{p['proj_fp_ros']:.1f}" if p.get('proj_fp_ros') is not None else "-"
            if sort_col == "fppg":
                result.append(
                    f"| {i} | {pname} | {p['fantasy_team']} | {p['nba_team']} "
                    f"| {p['fppg']:.2f} | {p['total_fp']:.1f} | {p['gp_pct']:.1f} "
                    f"| {mpg_str} | {p['eff_pct']:.1f} | {cv_str} | {iqr_str} "
                    f"| {proj_fp_str} | {p['proj_fppg_ros']:.2f} |"
                )
            else:
                result.append(
                    f"| {i} | {pname} | {p['fantasy_team']} | {p['nba_team']} "
                    f"| {p['total_fp']:.1f} | {p['fppg']:.2f} | {p['gp_pct']:.1f} "
                    f"| {mpg_str} | {p['eff_pct']:.1f} | {cv_str} | {iqr_str} "
                    f"| {proj_fp_str} | {p['proj_fppg_ros']:.2f} |"
                )
        result.append("")
        return result
    
    lines.extend(format_season_table("Season Best -> Total FP", sp.get("best_total_fp", []), "total_fp"))
    lines.extend(format_season_table("Season Best -> FPPG", sp.get("best_fppg", []), "fppg"))
    lines.extend(format_season_table("Season Worst -> Total FP", sp.get("worst_total_fp", []), "total_fp"))
    lines.extend(format_season_table("Season Worst -> FPPG", sp.get("worst_fppg", []), "fppg"))
    
    # Waiver pickups (weekly)
    waivers = bw.get("best_waivers", [])
    if waivers:
        lines.append("**Waiver Pickups (best_waivers):**")
        lines.append("| Player | NBA Team | Manager | GP | Total FP | FPPG |")
        lines.append("|--------|----------|---------|-----|----------|------|")
        for w in waivers:
            lines.append(f"| {w['player_name']} | {w['nba_team']} | {w['manager']} | {w['games']} | {w['total_fp']:.2f} | {w['fppg']:.2f} |")
        lines.append("")
    
    # Season Waiver Wire ROI -- REMOVED (rendered by Stats Corner viz)

    # Top free agents
    fa = bw.get("best_free_agents", [])[:5]
    if fa:
        lines.append("**Top Available Free Agents:**")
        lines.append("| Player | NBA Team | Position | Proj FPPG | Games This Week | Games Next Week |")
        lines.append("|--------|----------|----------|-----------|-----------------|-----------------|")
        for f in fa:
            pos_raw = f.get("positions", f.get("position", "?"))
            pos = ",".join(pos_raw) if isinstance(pos_raw, list) else str(pos_raw)
            gtw = f.get("games_this_week", "?")
            gnw = f.get("games_next_week", "?")
            lines.append(f"| {f['player_name']} | {f.get('nba_team', '?')} | {pos} | {f['projected_fppg']:.2f} | {gtw} | {gnw} |")
        lines.append("")
    
    # Bench Report -- REMOVED (rendered by Stats Corner viz)
    # Record Book -- REMOVED (rendered by Stats Corner viz)
    # Keeper Watch -- REMOVED (rendered by Stats Corner viz)
    # Draft Value Tracker -- REMOVED (rendered by Stats Corner viz)


    return "\n".join(lines)


def format_section_9_around_nba(data: dict, lookups: dict, injury_timelines: Optional[dict] = None, trades: Optional[list] = None, trades_context: Optional[dict] = None) -> str:
    """Format Section 9: Around the NBA (league context for web search)."""
    lines = ["## SECTION 9: AROUND THE NBA (League Context Only -> Web Search Deferred)\n"]
    
    # === TRADES (headline priority -> always first) ===
    if trades:
        num_trades = len(trades)
        lines.append(f"**TRADES THIS WEEK ({num_trades} trade{'s' if num_trades > 1 else ''} -> headline priority, cover FIRST):**")
        lines.append(f"NOTE: The first {num_trades} headline(s) in Around the NBA MUST cover {'these trades' if num_trades > 1 else 'this trade'}.")
        lines.append("Each trade headline paragraph MUST include a trade grade (letter grade A-F) for EACH manager involved,")
        lines.append("analyzing fairness considering player value, projected production, and draft pick compensation.")
        lines.append("")
        
        for i, trade in enumerate(trades, 1):
            lines.append(f"Trade {i}: {trade['manager_a']} -> {trade['manager_b']}")
            
            # Format sends_a
            sends_a_parts = []
            for item in trade["sends_a"]:
                if item["is_player"]:
                    if item.get("keeper_tier"):
                        sends_a_parts.append(
                            f"{item['item']} ({item['proj_fppg']:.1f} proj FPPG, "
                            f"{item['keepability_score']:.1f} keep score, {item['keeper_tier']})"
                        )
                    else:
                        sends_a_parts.append(f"{item['item']} ({item['proj_fppg']:.1f} proj FPPG)")
                else:
                    sends_a_parts.append(item["item"])
            lines.append(f"- {trade['manager_a']} sends: {' + '.join(sends_a_parts)}")
            
            # Format sends_b
            sends_b_parts = []
            for item in trade["sends_b"]:
                if item["is_player"]:
                    if item.get("keeper_tier"):
                        sends_b_parts.append(
                            f"{item['item']} ({item['proj_fppg']:.1f} proj FPPG, "
                            f"{item['keepability_score']:.1f} keep score, {item['keeper_tier']})"
                        )
                    else:
                        sends_b_parts.append(f"{item['item']} ({item['proj_fppg']:.1f} proj FPPG)")
                else:
                    sends_b_parts.append(item["item"])
            lines.append(f"- {trade['manager_b']} sends: {' + '.join(sends_b_parts)}")
            lines.append("")
        
        lines.append("---")
        lines.append("")
    
    # === SEASON TRADE LOG & PICK OWNERSHIP (from TRADES.json) ===
    if trades_context:
        season_trades = trades_context.get("season_trades", [])
        if season_trades:
            lines.append("**SEASON TRADE LOG (from TRADES.json -> DO NOT invent trade history beyond this list):**")
            for t in season_trades:
                week = t.get("week", "?")
                date = t.get("date", "?")
                side_a = t.get("side_a", {})
                side_b = t.get("side_b", {})
                mgr_a = side_a.get("manager", "?")
                mgr_b = side_b.get("manager", "?")
                
                a_parts = side_a.get("sent_players", []) + side_a.get("sent_picks", [])
                b_parts = side_b.get("sent_players", []) + side_b.get("sent_picks", [])
                
                lines.append(f"- Week {week} ({date}): {mgr_a} sends {' + '.join(a_parts)} <-> {mgr_b} sends {' + '.join(b_parts)}")
            lines.append(f"Total trades this season: {len(season_trades)}")
            lines.append("")
        
        # Draft pick ownership
        pick_summary = trades_context.get("draft_pick_summary", {})
        if any(v["owns"] or v["traded_away"] for v in pick_summary.values()):
            lines.append("**DRAFT PICK OWNERSHIP (non-default only -> if not listed, manager owns their own pick):**")
            for mgr in MANAGERS:
                summary = pick_summary.get(mgr, {})
                owns = summary.get("owns", [])
                traded = summary.get("traded_away", [])
                if owns or traded:
                    parts = []
                    if owns:
                        parts.append(f"Acquired: {', '.join(owns)}")
                    if traded:
                        parts.append(f"Traded away: {', '.join(traded)}")
                    lines.append(f"- {mgr}: {' | '.join(parts)}")
            lines.append("")
            lines.append("WARNING: Do NOT suggest a manager trades a pick they no longer own.")
            lines.append("WARNING: Do NOT claim managers are \'competing for the #1 pick\' without checking who owns whose picks.")
            lines.append("")
    
    # Rostered players
    rosters = data.get("rosters", {})
    lines.append("**Rostered Players by Manager:**")
    for mgr, players in rosters.items():
        lines.append(f"- {mgr} ({len(players)}): {', '.join(sorted(players))}")
    lines.append("")
    
    # Current injuries from team health, enriched with timelines
    lines.append("**Current Injuries (from current_team_health):**")
    for mgr, th in lookups["team_health"].items():
        for p in th.get("injured_players", []):
            status = "season-long" if p.get("is_season_long") else p.get("status", "out")
            notes = p.get("notes", "")
            remaining = p.get("remaining_weeks", 0)
            player_name = p["player"]
            
            # Enrich with timeline data if available
            timeline = (injury_timelines or {}).get(player_name)
            if timeline and timeline["weeks_missed"] > 0:
                missed = timeline["weeks_missed"]
                first_out = timeline["first_week_out"]
                total = timeline["total_projected"]
                
                if status == "season-long":
                    lines.append(f"- {player_name} ({mgr}): {notes}. Has missed {missed} weeks (since week {first_out}). Out for season.")
                elif status == "returning":
                    lines.append(f"- {player_name} ({mgr}): {notes}. Set to return after {missed} week(s) out (since week {first_out}).")
                else:
                    lines.append(f"- {player_name} ({mgr}): {notes}. Missed {missed} of ~{total} weeks so far (since week {first_out}). ~{remaining} weeks remaining.")
            else:
                lines.append(f"- {player_name} ({mgr}): {notes}. ~{remaining} weeks remaining. Status: {status}")
    lines.append("")
    
    # Suggested web search topics
    lines.append("**SUGGESTED WEB SEARCH TOPICS (for drafting pass):**")
    lines.append("1. [Search for injury updates on key injured players]")
    lines.append("2. [Search for trade deadline rumors if near deadline]")
    lines.append("3. [Search for recent performances of POTW candidate]")
    lines.append("4. [Search for any breaking news on rostered players]")
    lines.append("")
    lines.append("NOTE: Claude should perform web searches during the drafting pass to find current NBA news relevant to rostered players.")
    lines.append("")
    
    return "\n".join(lines)


def format_section_10_rumor_mill(data: dict, lookups: dict, trades_context: Optional[dict] = None) -> str:
    """Format Section 10: Rumor Mill."""
    lines = ["## SECTION 10: RUMOR MILL\n"]
    
    rm = data.get("rumor_mill", {})
    
    # === DRAFT PICK OWNERSHIP CONTEXT (for realistic trade ideas) ===
    if trades_context:
        pick_summary = trades_context.get("draft_pick_summary", {})
        if any(v["owns"] or v["traded_away"] for v in pick_summary.values()):
            lines.append("**DRAFT PICK CONTEXT (from TRADES.json -> trade ideas must respect current ownership):**")
            for mgr in MANAGERS:
                summary = pick_summary.get(mgr, {})
                owns = summary.get("owns", [])
                traded = summary.get("traded_away", [])
                if owns or traded:
                    parts = []
                    if owns:
                        parts.append(f"Extra picks: {', '.join(owns)}")
                    if traded:
                        parts.append(f"No longer owns: {', '.join(traded)}")
                    lines.append(f"- {mgr}: {' | '.join(parts)}")
            lines.append("")
    
    # Trade ideas
    trades = rm.get("trade_ideas", [])
    if trades:
        lines.append("**Trade Ideas:**")
        for i, t in enumerate(trades, 1):
            gives_list = t.get("gives_a", [])
            receives_list = t.get("receives_a", [])
            # Clean up items that start with + (like "+ a high pick")
            gives_clean = [g.lstrip("+ ") if g.startswith("+") else g for g in gives_list]
            receives_clean = [r.lstrip("+ ") if r.startswith("+") else r for r in receives_list]
            gives = " + ".join(gives_clean) if gives_clean else "?"
            receives = " + ".join(receives_clean) if receives_clean else "?"
            lines.append(f"{i}. {t['manager_a']} sends {gives} <-> {t['manager_b']} sends {receives}")
            lines.append(f"   - Type: {t.get('trade_type', 'swap')}")
            lines.append(f"   - Rationale: \"{t.get('rationale', '')}\"")
            lines.append(f"   - Fit score: {t.get('fit_score', 0):.2f}")
        lines.append("")
    
    # Free agent targets
    fa_targets = rm.get("free_agent_targets", [])
    if fa_targets:
        lines.append("**Free Agent Targets:**")
        for f in fa_targets:
            pos = ",".join(f.get("positions", []))
            lines.append(f"- {f['player_name']} ({pos}) -> {f['target_manager']}: {f['projected_fppg']:.2f} proj FPPG. \"{f.get('reason', '')}\"")
        lines.append("")
    
    # Hot streaks
    hot = rm.get("hot_streak_candidates", [])
    if hot:
        lines.append("**Hot Streaks:**")
        for i, h in enumerate(hot, 1):
            overperf = h.get("overperformance_index", 0)
            overperf_14 = h.get("overperformance_index_last_14_days")
            overperf_14_str = f"{overperf_14:.1f}%" if overperf_14 is not None else "N/A"
            note = h.get("trade_value_note", "")
            lines.append(f"{i}. {h['player_name']} ({h['manager']}): +{overperf:.1f}% above proj. Last 14d: +{overperf_14_str}. NOTE: \"{note}\"")
        lines.append("")
    
    # Slump watch / drop candidates
    drops = rm.get("drop_candidates", [])
    if drops:
        lines.append("**Slump Watch:**")
        for i, d in enumerate(drops, 1):
            underperf = abs(d.get("underperformance_index", 0))
            underperf_14 = d.get("underperformance_index_last_14_days")
            underperf_14_str = f"{abs(underperf_14):.1f}%" if underperf_14 is not None else "N/A"
            better_fa = d.get("better_fa_available") or "None"
            lines.append(f"{i}. {d['player_name']} ({d['manager']}): -{underperf:.1f}% below proj. Last 14d: -{underperf_14_str}. Better FA: {better_fa}")
        lines.append("")
    
    # Season injury burden for trade context
    lines.append("**Season Injury Burden (for trade context):**")
    lines.append("| Manager | Total Burden % | Total Injury Games | Games Lost to Injury (non-IL) | IL Games | Top IL Player |")
    lines.append("|---------|---------------|-------------------|-------------------------------|----------|---------------|")
    sorted_ib = sorted(
        lookups["injury_burden"].items(),
        key=lambda x: x[1].get("total_injury_burden_pct", 0),
        reverse=True
    )
    for mgr, ib in sorted_ib:
        top_il = ib["il_players"][0] if ib.get("il_players") else {"player": "None", "games": 0}
        lines.append(f"| {mgr} | {ib.get('total_injury_burden_pct', 0):.1f}% | {ib.get('total_injury_games', 0)} | {ib.get('non_il_injury_games', 0)} | {ib.get('il_injury_games', 0)} | {top_il['player']} ({top_il['games']}) |")
    lines.append("")
    
    # Trade history -- build_all_time_records() can return None.
    atr = data.get("all_time_records") or {}
    partners = atr.get("trade_partners") or {}
    if partners:
        lines.append("**Trade History (all_time_records.trade_partners):**")
        lines.append("| Trade Partners | All-Time Deals |")
        lines.append("|----------------|---------------|")
        sorted_partners = sorted(partners.items(), key=lambda x: x[1], reverse=True)
        for pair, count in sorted_partners:
            # Format pair name nicely
            pair_formatted = pair.replace("_and_", " & ")
            lines.append(f"| {pair_formatted} | {count} |")
        lines.append("")
    
    return "\n".join(lines)


# =============================================================================
# MAIN FORMATTER
# =============================================================================

def compute_storyline_alerts(data: dict, lookups: dict, predictions: Optional[dict] = None) -> list[str]:
    """
    Pre-compute narrative conclusions so Claude doesn't have to deduce them.
    Returns a list of explicit storyline statements ready to be woven into prose.
    
    Args:
        data: The stats report JSON
        lookups: Cross-reference lookup tables
        predictions: Optional dict of last week's betting line predictions for upset detection
    """
    alerts = []
    # Defensive: build_all_time_records / build_current_streaks / build_metadata could be missing or None.
    week = (data.get("metadata") or {}).get("week", 0)
    streaks = data.get("current_streaks") or {}
    atr = data.get("all_time_records") or {}
    all_time_win_streaks = atr.get("longest_win_streaks") or {}
    all_time_loss_streaks = atr.get("longest_loss_streaks") or {}
    career_standings = {c["manager"]: c for c in (atr.get("career_standings") or [])}
    
    # Track what we've already mentioned to avoid duplicates
    mentioned_managers = set()
    
    # === RECORD UPDATES (from record_updates field) ===
    for update in data.get("record_updates", []):
        desc = update.get("description", "")
        val = update.get("value", "")
        update_type = update.get("type", "")
        
        if "highest_weekly_team_score" in update_type:
            # Find who scored it
            for ms in (data.get("matchup_summaries") or []):
                if ms["stats_a"]["total_fp"] == val:
                    alerts.append(f" -> SEASON RECORD: {ms['manager_a']} scored {val:.2f} FP -> the highest team score of the {CURRENT_SEASON} season")
                    break
                elif ms["stats_b"]["total_fp"] == val:
                    alerts.append(f" -> SEASON RECORD: {ms['manager_b']} scored {val:.2f} FP -> the highest team score of the {CURRENT_SEASON} season")
                    break
        elif "win_streak" in update_type:
            if "new" in desc.lower() or "sets" in desc.lower():
                # Parse who set it
                for mgr in MANAGERS:
                    if mgr.lower() in desc.lower():
                        current = streaks.get(mgr, {}).get("win_streak", 0)
                        all_time = all_time_win_streaks.get(mgr, {}).get("length", 0)
                        all_time_season = all_time_win_streaks.get(mgr, {}).get("season", "")
                        
                        if current > all_time:
                            alerts.append(f" -> ALL-TIME RECORD: {mgr} set a new personal best with a {current}-game win streak, surpassing their previous record of {all_time} ({all_time_season})")
                        else:
                            # Season record but not all-time - include context about all-time
                            gap = all_time - current
                            if gap == 1:
                                alerts.append(f" -> SEASON RECORD: {mgr} set a new season-best {current}-game win streak -> just 1 win from tying their all-time record of {all_time} ({all_time_season})")
                            elif gap == 2:
                                alerts.append(f" -> SEASON RECORD: {mgr} set a new season-best {current}-game win streak -> 2 wins from tying their all-time record of {all_time} ({all_time_season})")
                            else:
                                alerts.append(f" -> SEASON RECORD: {mgr} set a new season-best {current}-game win streak (all-time record: {all_time})")
                        mentioned_managers.add(mgr)
                        break
        elif "loss_streak" in update_type:
            for mgr in MANAGERS:
                if mgr.lower() in desc.lower():
                    current = streaks.get(mgr, {}).get("loss_streak", 0)
                    alerts.append(f" -> LOSING STREAK: {mgr} has now lost {current} straight")
                    break
    
    # === LOSING STREAK ENDINGS ===
    for matchup in data.get("matchup_summaries", []):
        winner = matchup.get("winner")
        loser = matchup.get("manager_a") if winner == matchup.get("manager_b") else matchup.get("manager_b")
        margin = matchup.get("margin", 0)
        
        winner_streaks = streaks.get(winner, {})
        season_longest_loss = winner_streaks.get("season_longest_loss_streak", 0)
        season_longest_loss_weeks = winner_streaks.get("season_longest_loss_weeks", "")
        current_win = winner_streaks.get("win_streak", 0)
        
        # If they have a win streak of 1 and had a significant losing streak that just ended
        if current_win == 1 and season_longest_loss >= 3 and season_longest_loss_weeks:
            try:
                parts = season_longest_loss_weeks.split("-")
                if len(parts) == 2:
                    loss_start, loss_end = int(parts[0]), int(parts[1])
                    if loss_end == week - 1:  # Lost through last week, won this week
                        alerts.append(f" -> STREAK SNAPPED: {winner} ends {season_longest_loss}-game losing streak (weeks {loss_start}-{loss_end}) with Week {week} victory over {loser}, {matchup.get('stats_a' if winner == matchup['manager_a'] else 'stats_b', {}).get('total_fp', 0):.2f} -> {matchup.get('stats_b' if winner == matchup['manager_a'] else 'stats_a', {}).get('total_fp', 0):.2f}")
            except (ValueError, TypeError):
                pass
    
    # === UPSET DETECTION (from last week's betting lines) ===
    if predictions:
        for matchup in data.get("matchup_summaries", []):
            winner = matchup.get("winner")
            loser = matchup.get("manager_a") if winner == matchup.get("manager_b") else matchup.get("manager_b")
            mgr_a = matchup.get("manager_a")
            mgr_b = matchup.get("manager_b")
            margin = matchup.get("margin", 0)
            
            # Look up last week's prediction for this matchup
            key = tuple(sorted([mgr_a, mgr_b]))
            pred = predictions.get(key)
            
            if pred:
                favorite = pred.get("favorite")
                underdog = pred.get("underdog")
                fav_prob = pred.get("win_prob_favorite", 50)
                dog_prob = pred.get("win_prob_underdog", 50)
                spread = pred.get("spread", 0)
                
                # Upset = underdog won
                if winner == underdog and fav_prob >= 51:  # Flag any upset where favorite had majority probability
                    # Calculate how much of an upset it was
                    upset_magnitude = fav_prob - 50  # How favored was the favorite?
                    
                    # Get actual scores
                    winner_score = matchup.get('stats_a' if winner == mgr_a else 'stats_b', {}).get('total_fp', 0)
                    loser_score = matchup.get('stats_b' if winner == mgr_a else 'stats_a', {}).get('total_fp', 0)
                    
                    if fav_prob >= 70:
                        alerts.append(f" -> MAJOR UPSET: {winner} ({dog_prob:.0f}% win prob) stuns {loser} ({fav_prob:.0f}% favorite), {winner_score:.2f} -> {loser_score:.2f}")
                    elif fav_prob >= 60:
                        alerts.append(f" -> UPSET: {winner} ({dog_prob:.0f}% win prob) defeats favored {loser} ({fav_prob:.0f}%), {winner_score:.2f} -> {loser_score:.2f}")
                    else:
                        alerts.append(f" -> UPSET: {winner} wins as {dog_prob:.0f}% underdog over {loser} ({fav_prob:.0f}% favorite)")
                
                # Also flag when favorite wins but covers big spread
                elif winner == favorite and spread and abs(spread) >= 30:
                    # Did they cover the spread?
                    if margin > abs(spread):
                        alerts.append(f" -> SPREAD COVERED: {winner} beat the {abs(spread):.1f}-point spread with {margin:.2f}-point victory")
                    elif margin < abs(spread) - 50:
                        # They won but massively underperformed expectations
                        alerts.append(f" -> CLOSE CALL: {winner} was favored by {abs(spread):.1f} but only won by {margin:.2f}")
    
    
    # === CAREER MILESTONES ===
    for mgr, career in career_standings.items():
        wins = career.get("wins", 0)
        
        # Check if they won this week
        won_this_week = False
        opponent = None
        for matchup in data.get("matchup_summaries", []):
            if matchup.get("winner") == mgr:
                won_this_week = True
                opponent = matchup.get("manager_a") if mgr == matchup.get("manager_b") else matchup.get("manager_b")
                break
        
        if not won_this_week:
            continue
        
        # Round number milestones
        if wins in [50, 75, 100, 125, 150, 175, 200]:
            mgr_streaks = streaks.get(mgr, {})
            loss_streak = mgr_streaks.get("season_longest_loss_streak", 0)
            loss_weeks = mgr_streaks.get("season_longest_loss_weeks", "")
            
            # Figure out how long they were waiting
            wait_info = ""
            if loss_weeks and loss_streak >= 2:
                try:
                    parts = loss_weeks.split("-")
                    if len(parts) == 2:
                        loss_start, loss_end = int(parts[0]), int(parts[1])
                        if loss_end == week - 1:
                            weeks_waiting = loss_streak
                            wait_info = f" after a {weeks_waiting}-week wait (stuck at {wins-1} since Week {loss_start})"
                except (ValueError, TypeError):
                    pass
            
            alerts.append(f" -> MILESTONE: {mgr} secures career win #{wins}{wait_info}")
    
    # === WIN STREAK CONTEXT (approaching records) ===
    for mgr, mgr_streaks in streaks.items():
        if mgr in mentioned_managers:
            continue  # Already covered in record updates
            
        current_win = mgr_streaks.get("win_streak", 0)
        if current_win >= 3:
            all_time = all_time_win_streaks.get(mgr, {}).get("length", 0)
            all_time_season = all_time_win_streaks.get(mgr, {}).get("season", "")
            season_best = mgr_streaks.get("season_longest_win_streak", 0)
            
            if current_win > all_time:
                alerts.append(f" -> ALL-TIME RECORD: {mgr}'s {current_win}-game win streak is a new personal best (previous: {all_time} in {all_time_season})")
            elif current_win == all_time:
                alerts.append(f" -> RECORD TIED: {mgr}'s {current_win}-game win streak ties their all-time best ({all_time_season})")
            elif current_win == all_time - 1:
                alerts.append(f" -> RECORD WATCH: {mgr} has won {current_win} straight -> 1 win from tying their all-time record of {all_time} ({all_time_season})")
            elif current_win == all_time - 2 and current_win >= 4:
                alerts.append(f" -> STREAK ALERT: {mgr} has won {current_win} straight (all-time record: {all_time})")
    
    # === MATCHUP CONTEXT (blowouts, close games, upsets) ===
    biggest_blowout = atr.get("biggest_blowout", {})
    closest_game = atr.get("closest_game", {})
    
    for matchup in data.get("matchup_summaries", []):
        margin = matchup.get("margin", 0)
        winner = matchup.get("winner")
        loser = matchup.get("manager_a") if winner == matchup.get("manager_b") else matchup.get("manager_b")
        
        # Check for historic margins
        if margin > 500:
            if biggest_blowout.get("margin", 0) and margin > biggest_blowout["margin"]:
                alerts.append(f" -> ALL-TIME BLOWOUT: {winner}'s {margin:.2f}-point victory over {loser} is the largest margin in league history!")
            else:
                alerts.append(f" -> BLOWOUT: {winner} crushed {loser} by {margin:.2f} points")
        elif margin < 10:
            if closest_game.get("margin", float('inf')) and margin < closest_game["margin"]:
                alerts.append(f" -> ALL-TIME CLOSEST: {winner}'s {margin:.2f}-point victory over {loser} is the closest game in league history!")
            else:
                alerts.append(f" -> NAIL-BITER: {winner} edged {loser} by just {margin:.2f} points")
        elif margin < 25:
            alerts.append(f" -> CLOSE GAME: {winner} held off {loser} by {margin:.2f} points")
    
    # === STANDINGS IMPLICATIONS ===
    standings = data.get("current_standings", [])
    if standings:
        leader = standings[0]
        second = standings[1] if len(standings) > 1 else None
        
        if leader and second:
            leader_record = leader.get("record", "0-0").split("-")
            second_record = second.get("record", "0-0").split("-")
            try:
                leader_wins = int(leader_record[0])
                second_wins = int(second_record[0])
                games_back = leader_wins - second_wins
                
                if games_back >= 4:
                    alerts.append(f" -> STANDINGS: {leader['manager']} leads by {games_back} games over {second['manager']} ({leader['record']} vs {second['record']})")
            except (ValueError, IndexError):
                pass
    
    # === MAGIC NUMBER ===
    # power_rankings is None in --fast mode.
    for pr in (data.get("power_rankings") or []):
        if pr.get("magic_number") and pr.get("rank") == 1:
            alerts.append(f" -> CLINCH WATCH: {pr['manager']}'s magic number is {pr['magic_number']}")
    
    return alerts


def format_stats_report(data: dict, predictions: Optional[dict] = None, prev_grades: Optional[dict] = None, player_projs: Optional[dict] = None, potw_history: Optional[list] = None, potw_career_stats: Optional[dict] = None, prev_title_odds: Optional[dict] = None, injury_timelines: Optional[dict] = None, trades: Optional[list] = None, trades_context: Optional[dict] = None, commissioner_notes: Optional[str] = None) -> str:
    """
    Format the complete stats report as newsletter-ready markdown.
    
    Args:
        data: The stats report JSON
        predictions: Optional dict of last week's betting line predictions for upset detection
        prev_grades: Optional dict of last week's report card grades {manager: grade}
        player_projs: Optional dict of player projected FPPG {player_name: fppg}
        potw_history: Optional list of previous POTW winners this season
        potw_career_stats: Optional dict of career POTW stats across all seasons
        prev_title_odds: Optional dict of last week's title odds {manager: pct}
    """
    lookups = build_lookup_tables(data)
    
    # Header
    meta = data.get("metadata", {})
    week = meta.get("week", "?")
    season = meta.get("season_year", "?")
    date_range = meta.get("date_range", {})
    start = date_range.get("start", "?")
    end = date_range.get("end", "?")
    generated = meta.get("generated_at", datetime.now().isoformat())
    
    header = f"""# WEEK {week} STATS REPORT -> NEWSLETTER INPUT
# Season {season} | Week {week} ({start} -> {end})
# Generated: {generated}
# 
# This file contains all data needed for the newsletter, organized by section.
# Claude's job: Write prose using this data. Do NOT extract -> it's already done.
# For Section 9 (Around the NBA): Perform web searches for current news.

---

"""
    
    # Compute storyline alerts - pre-deduced narrative conclusions
    alerts = compute_storyline_alerts(data, lookups, predictions)
    
    storylines_section = ""
    if alerts:
        storylines_section = "## -> STORYLINE ALERTS -> READ FIRST -> \n\n"
        storylines_section += "**These are the key narratives this week. Do NOT contradict these facts:**\n\n"
        for alert in alerts:
            storylines_section += f"- {alert}\n"
        storylines_section += "\n---\n\n"
    
    # Commissioner notes (from weeklycontextinput_weekN.json -> "notes" field)
    commissioner_section = ""
    if commissioner_notes:
        commissioner_section = "## -> COMMISSIONER NOTE ->\n\n"
        commissioner_section += f"{commissioner_notes}\n\n---\n\n"
    
    sections = [
        format_section_1_matchup_summaries(data, lookups),
        format_section_2_report_cards(data, lookups, prev_grades),
        format_section_3_betting_lines(data, lookups, player_projs),
        format_section_4_potw(data, lookups, potw_history, potw_career_stats),
        format_section_5_fun_facts(data, lookups),
        format_section_6_what_if(data, lookups),
        format_section_7_power_rankings(data, lookups, prev_title_odds),
        format_section_8_stats_corner(data, lookups),
        format_section_9_around_nba(data, lookups, injury_timelines, trades, trades_context),
        format_section_10_rumor_mill(data, lookups, trades_context),
    ]
    
    return header + storylines_section + commissioner_section + "\n".join(sections)


def main():
    parser = argparse.ArgumentParser(
        description="Convert stats report JSON to newsletter-ready markdown"
    )
    parser.add_argument(
        "--week",
        type=int,
        help="Week number (looks for output/stats_report_weekN.json)",
    )
    parser.add_argument(
        "--input",
        type=str,
        help="Input JSON file path (alternative to --week)",
    )
    parser.add_argument(
        "--output",
        type=str,
        help="Output markdown file path (default: same as input with .md extension)",
    )
    parser.add_argument(
        "--base-path",
        type=str,
        default=".",
        help="Base path to project directory (default: current directory)",
    )
    parser.add_argument(
        "--allow-missing-projections",
        action="store_true",
        help="Allow running without PLAYERLIST.xlsx projections (use only for debugging; output may be incomplete)",
    )
    parser.add_argument(
        "--last-week-recap",
        type=str,
        help="Path to LAST_WEEK_RECAP.md for upset detection (default: config/LAST_WEEK_RECAP.md)",
    )
    
    args = parser.parse_args()
    
    # Determine input path
    if args.input:
        input_path = Path(args.input)
    elif args.week:
        input_path = Path(args.base_path) / f"output/stats_report_week{args.week}.json"
    else:
        parser.error("Must specify either --week or --input")
    
    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        return 1
    
    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = input_path.with_suffix(".md")
    
    # Load current week's data first to get week number
    print(f"Loading: {input_path}")
    data = load_stats_report(input_path)
    current_week = data.get("metadata", {}).get("week", 0)
    
    # Load betting predictions from PREVIOUS week's JSON (more reliable than parsing recap)
    predictions = None
    output_dir = Path(args.base_path) / "output"  # Per PROJECTSTRUCTURE.md
    
    if current_week > 1:
        predictions = load_last_week_betting_predictions(current_week, output_dir)
        if predictions:
            print(f"Loaded betting predictions from week {current_week - 1} JSON ({len(predictions)} matchup(s))")
        else:
            # Fall back to parsing LAST_WEEK_RECAP.md
            if args.last_week_recap:
                recap_path = Path(args.last_week_recap)
            else:
                recap_path = Path(args.base_path) / "config" / "LAST_WEEK_RECAP.md"
            
            if recap_path.exists():
                print(f"Falling back to parsing predictions from: {recap_path}")
                predictions = parse_last_week_predictions(recap_path)
                if predictions:
                    print(f"  Found {len(predictions)} matchup prediction(s)")
                else:
                    print("  No betting line predictions found in recap file")
            else:
                print(f"Note: No previous week's JSON or LAST_WEEK_RECAP.md found -> skipping upset detection")
    
    # Load previous week's report card grades
    prev_grades = None
    if current_week > 1:
        prev_grades = load_last_week_grades(current_week, output_dir)
        if prev_grades:
            print(f"Loaded previous grades from week {current_week - 1}: {', '.join(f'{m} {g}' for m, g in prev_grades.items())}")
    
    # Load previous week's title odds for delta computation
    prev_title_odds = None
    if current_week > 1:
        prev_title_odds = load_last_week_title_odds(current_week, output_dir)
        if prev_title_odds:
            print(f"Loaded previous title odds from week {current_week - 1}: {', '.join(f'{m} {o:.1f}%' for m, o in prev_title_odds.items())}")
    
    # Load player projections (required unless --allow-missing-projections)
    player_projs = load_player_projections(Path(args.base_path))

    # An all-zero projection table is the legitimate final-week state (no
    # upcoming games left to project), NOT a missing/unreadable file. Treat it
    # like the allow-missing fallback so the report still generates instead of
    # hard-failing on the last week of the season.
    projections_all_zero = bool(player_projs) and all(v == 0 for v in player_projs.values())
    if projections_all_zero:
        print("NOTE: PLAYERLIST.xlsx projections are all zero (final week -- nothing left to project). Continuing with JSON fallbacks.")
        player_projs = {}

    if player_projs:
        print(f"Loaded projections for {len(player_projs)} players from PLAYERLIST.xlsx")
    elif args.allow_missing_projections or projections_all_zero:
        print("WARNING: Proceeding without PLAYERLIST.xlsx projections -> using JSON key_player fallback (some projection-based fields may be incomplete)")
    else:
        missing_reason = "openpyxl is not installed" if openpyxl is None else "PLAYERLIST.xlsx missing or unreadable"
        print(f"ERROR: Projections are required, but they could not be loaded ({missing_reason}).")
        print("Fix: ensure data/PLAYERLIST.xlsx exists and is readable, and that openpyxl is installed (pip install openpyxl).")
        print("If you *really* need to run anyway (debugging only), pass --allow-missing-projections.")
        return 1
    
    # Load injury timelines from INJURY_OVERRIDES.json
    injury_timelines = load_injury_timelines(Path(args.base_path), current_week)
    if injury_timelines:
        print(f"Loaded injury timelines for {len(injury_timelines)} players from INJURY_OVERRIDES.json")
    
    # Load trades from weekly context input (with keepability context if available)
    keeper_watch_players = data.get("keeper_watch", {}).get("players", [])
    trades = load_weekly_trades(Path(args.base_path), current_week, player_projs or {}, keeper_watch_players=keeper_watch_players)
    if trades:
        print(f"Loaded {len(trades)} trade(s) from weekly context input")

    # Auto-sync weekly context trades into TRADES.json
    synced = sync_weekly_trades_to_trades_json(Path(args.base_path), current_week)
    if synced > 0:
        print(f"Synced {synced} new trade(s) to TRADES.json")

    # Sync trade partner counts to RECORDS.json (historical + current season)
    trade_partners = sync_trade_partners_to_records(Path(args.base_path))
    if trade_partners:
        print(f"Synced trade partner counts to RECORDS.json: {len(trade_partners)} partner pair(s)")

    # Load season trade log and draft pick ownership from TRADES.json
    trades_context = load_trades_context(Path(args.base_path), current_week)
    if trades_context:
        n_season = len(trades_context.get("season_trades", []))
        n_moved = sum(
            len(v["owns"]) for v in trades_context.get("draft_pick_summary", {}).values()
        )
        print(f"Loaded TRADES.json: {n_season} season trade(s), {n_moved} pick(s) changed hands")
    
    # Load POTW history from config file + previous weeks' JSONs
    season = data.get("metadata", {}).get("season_year", "")
    potw_history, potw_all_data = load_potw_history(current_week, Path(args.base_path), season)
    potw_career_stats = compute_potw_career_stats(potw_all_data) if potw_all_data else None
    if potw_history:
        print(f"Loaded POTW history: {len(potw_history)} previous winner(s) this season")
    
    # Load commissioner notes from weekly context input
    commissioner_notes = load_weekly_context_notes(Path(args.base_path), current_week)
    if commissioner_notes:
        print(f"Loaded commissioner note from weeklycontextinput_week{current_week}.json")

    print(f"Formatting Week {current_week} report...")
    markdown = format_stats_report(data, predictions, prev_grades, player_projs, potw_history, potw_career_stats, prev_title_odds, injury_timelines, trades, trades_context, commissioner_notes)
    
    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(markdown)
    
    print(f"Saved: {output_path}")
    print(f"Lines: {len(markdown.splitlines())}")
    
    # Save current week's POTW winner to history file
    winner = (data.get("player_of_week") or {}).get("winner") or {}
    if winner and season:
        saved = save_potw_winner(Path(args.base_path), season, winner, current_week)
        if saved:
            print(f"Saved POTW winner to POTW_HISTORY.json: Week {current_week} -> {winner['player_name']}")
        else:
            print(f"POTW history: Week {current_week} already recorded (skipped)")
    
    return 0


if __name__ == "__main__":
    exit(main())


def _format_record_book_expanded(rb_data: dict) -> list:
    """
    Format the expanded Record Book with 6-category leaderboard tables.

    Outputs clear markdown tables for each record category so the LLM
    newsletter drafter has full top-10 data for rich narratives.

    Categories: Team Records, Player Records, Rookie Records,
    Draft & Trades, Manager Records, Manager Milestones.

    Current-season entries are marked with * suffix.
    """
    lines = []

    # ---- TEAM RECORDS ----
    team_records = rb_data.get("team_records", [])
    if team_records:
        lines.append("**Record Book -- Team Records:**")
        lines.append("")

        for rec in team_records:
            record_name = rec.get("record", "")
            entries = rec.get("entries", [])
            is_new = rec.get("is_new_record", False)
            new_tag = " [NEW RECORD!]" if is_new else ""

            lines.append(f"*{record_name}{new_tag}*")

            if not entries:
                lines.append("*Awaiting historical data*")
                lines.append("")
                continue

            lines.append("| Rank | Value | Manager | Season | Detail |")
            lines.append("|------|-------|---------|--------|--------|")

            for e in entries:
                rank = e.get("rank", "")
                value = e.get("value", "")
                holder = e.get("holder", "")
                season = e.get("season", "")
                detail = e.get("detail", "")
                current_mark = "*" if e.get("is_current_season") else ""

                # Format value
                if isinstance(value, float):
                    if abs(value) < 10:
                        val_str = f"{value:.2f}"
                    else:
                        val_str = f"{value:,.1f}"
                else:
                    val_str = str(value)

                lines.append(
                    f"| {rank} "
                    f"| {val_str}{current_mark} "
                    f"| {holder} "
                    f"| {season} "
                    f"| {detail} |"
                )

            lines.append("")

    # ---- PLAYER RECORDS ----
    player_records = rb_data.get("player_records", [])
    if player_records:
        lines.append("**Record Book -- Player Records:**")
        lines.append("")

        for rec in player_records:
            record_name = rec.get("record", "")
            entries = rec.get("entries", [])
            is_new = rec.get("is_new_record", False)
            new_tag = " [NEW RECORD!]" if is_new else ""
            min_gp = rec.get("min_gp")
            gp_note = f" (min {min_gp} GP)" if min_gp else ""

            lines.append(f"*{record_name}{gp_note}{new_tag}*")

            if not entries:
                lines.append("*Awaiting historical data*")
                lines.append("")
                continue

            lines.append("| Rank | Value | Player | Manager | Season | Detail |")
            lines.append("|------|-------|--------|---------|--------|--------|")

            for e in entries:
                rank = e.get("rank", "")
                value = e.get("value", "")
                player = e.get("player", "")
                manager = e.get("manager", "")
                season = e.get("season", "")
                detail = e.get("detail", "")
                current_mark = "*" if e.get("is_current_season") else ""

                if isinstance(value, float):
                    if abs(value) < 10:
                        val_str = f"{value:.2f}"
                    else:
                        val_str = f"{value:,.1f}"
                else:
                    val_str = str(value)

                lines.append(
                    f"| {rank} "
                    f"| {val_str}{current_mark} "
                    f"| {player} "
                    f"| {manager} "
                    f"| {season} "
                    f"| {detail} |"
                )

            lines.append("")

    # ---- ROOKIE RECORDS ----
    rookie_records = rb_data.get("rookie_records", [])
    if rookie_records:
        lines.append("**Record Book -- Rookie Records:**")
        lines.append("")

        for rec in rookie_records:
            record_name = rec.get("record", "")
            entries = rec.get("entries", [])
            is_new = rec.get("is_new_record", False)
            new_tag = " [NEW RECORD!]" if is_new else ""

            lines.append(f"*{record_name}{new_tag}*")

            if not entries:
                lines.append("*Awaiting historical data*")
                lines.append("")
                continue

            lines.append("| Rank | Value | Player | Manager | Season | Detail |")
            lines.append("|------|-------|--------|---------|--------|--------|")

            for e in entries:
                rank = e.get("rank", "")
                value = e.get("value", "")
                player = e.get("player", "")
                manager = e.get("manager", "")
                season = e.get("season", "")
                detail = e.get("detail", "")
                current_mark = "*" if e.get("is_current_season") else ""

                if isinstance(value, float):
                    if abs(value) < 10:
                        val_str = f"{value:.2f}"
                    else:
                        val_str = f"{value:,.1f}"
                else:
                    val_str = str(value)

                lines.append(
                    f"| {rank} "
                    f"| {val_str}{current_mark} "
                    f"| {player} "
                    f"| {manager} "
                    f"| {season} "
                    f"| {detail} |"
                )

            lines.append("")

    # ---- DRAFT & TRADES RECORDS ----
    draft_records = rb_data.get("draft_records", [])
    if draft_records:
        lines.append("**Record Book -- Draft & Trades:**")
        lines.append("")

        for rec in draft_records:
            record_name = rec.get("record", "")
            entries = rec.get("entries", [])
            is_new = rec.get("is_new_record", False)
            new_tag = " [NEW RECORD!]" if is_new else ""
            min_gp = rec.get("min_gp")
            gp_note = f" (min {min_gp} GP)" if min_gp else ""

            lines.append(f"*{record_name}{gp_note}{new_tag}*")

            if not entries:
                lines.append("*Awaiting historical data*")
                lines.append("")
                continue

            lines.append("| Rank | Value | Player | Manager | Season | Detail |")
            lines.append("|------|-------|--------|---------|--------|--------|")

            for e in entries:
                rank = e.get("rank", "")
                value = e.get("value", "")
                player = e.get("player", "")
                manager = e.get("manager", "")
                season = e.get("season", "")
                detail = e.get("detail", "")
                current_mark = "*" if e.get("is_current_season") else ""

                if isinstance(value, float):
                    if abs(value) < 10:
                        val_str = f"{value:.2f}"
                    else:
                        val_str = f"{value:,.1f}"
                else:
                    val_str = str(value)

                lines.append(
                    f"| {rank} "
                    f"| {val_str}{current_mark} "
                    f"| {player} "
                    f"| {manager} "
                    f"| {season} "
                    f"| {detail} |"
                )

            lines.append("")

    # ---- MANAGER RECORDS ----
    manager_records = rb_data.get("manager_records", [])
    if manager_records:
        lines.append("**Record Book -- Manager Records:**")
        lines.append("")

        for rec in manager_records:
            record_name = rec.get("record", "")
            entries = rec.get("entries", [])
            is_new = rec.get("is_new_record", False)
            new_tag = " [NEW RECORD!]" if is_new else ""

            lines.append(f"*{record_name}{new_tag}*")

            if not entries:
                lines.append("*Awaiting historical data*")
                lines.append("")
                continue

            lines.append("| Rank | Value | Player | Manager | Season | Detail |")
            lines.append("|------|-------|--------|---------|--------|--------|")

            for e in entries:
                rank = e.get("rank", "")
                value = e.get("value", "")
                player = e.get("player", "")
                manager = e.get("manager", "")
                season = e.get("season", "")
                detail = e.get("detail", "")

                if isinstance(value, float):
                    if abs(value) < 10:
                        val_str = f"{value:.2f}"
                    else:
                        val_str = f"{value:,.1f}"
                else:
                    val_str = str(value)

                lines.append(
                    f"| {rank} "
                    f"| {val_str} "
                    f"| {player} "
                    f"| {manager} "
                    f"| {season} "
                    f"| {detail} |"
                )

            lines.append("")

    # ---- MANAGER MILESTONES ----
    # Two trophy columns: "Titles" = regular-season championship (best record);
    # "Playoff Titles" = won the 2-week playoff bracket. The league tracks these
    # separately. Historical playoff_titles may be placeholder 0s pending
    # commissioner backfill.
    milestones = rb_data.get("manager_milestones", [])
    if milestones:
        lines.append("**Record Book -- Manager Milestones:**")
        lines.append("| Rank | Manager | Wins | Losses | Win% | Career Points | Titles | Playoff Titles |")
        lines.append("|------|---------|------|--------|------|---------------|--------|----------------|")

        for i, m in enumerate(milestones, 1):
            lines.append(
                f"| {i} "
                f"| {m.get('manager', '')} "
                f"| {m.get('wins', 0)} "
                f"| {m.get('losses', 0)} "
                f"| {m.get('win_pct', 0)} "
                f"| {m.get('career_points', 0):,.1f} "
                f"| {m.get('titles', 0)} "
                f"| {m.get('playoff_titles', 0)} |"
            )
        lines.append("")

    return lines

