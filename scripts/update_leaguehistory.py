import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# Add project root to path for config imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from modules.data_loader import CURRENT_SEASON_LONG

PLAYERLOG_XLSX_PATH = "data/PLAYERLOG.xlsx"
LINEUPS_XLSX_PATH = "data/LINEUPS.xlsx"
LEAGUEHISTORY_XLSX_PATH = "data/LEAGUEHISTORY.xlsx"
APPLIED_WEEKS_LEDGER_PATH = "config/.leaguehistory_applied_weeks.json"


def _load_applied_weeks_ledger(base_path: Path, season: str) -> dict:
    """Load the applied-weeks ledger. Returns dict keyed by season."""
    ledger_path = base_path / APPLIED_WEEKS_LEDGER_PATH
    if not ledger_path.exists():
        return {}
    try:
        with open(ledger_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}


def _save_applied_weeks_ledger(base_path: Path, ledger: dict) -> None:
    """Save the applied-weeks ledger atomically."""
    ledger_path = base_path / APPLIED_WEEKS_LEDGER_PATH
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    with open(ledger_path, "w", encoding="utf-8") as f:
        json.dump(ledger, f, indent=2)


def _is_week_already_applied(ledger: dict, season: str, week: int) -> bool:
    """Check whether (season, week) has already been accumulated."""
    season_entry = ledger.get(season) or {}
    applied = season_entry.get("applied_weeks", [])
    return week in applied


def _record_week_applied(ledger: dict, season: str, week: int) -> dict:
    """Record (season, week) in the ledger and return the updated ledger."""
    season_entry = ledger.setdefault(season, {"applied_weeks": [], "last_applied": None})
    if week not in season_entry["applied_weeks"]:
        season_entry["applied_weeks"].append(week)
        season_entry["applied_weeks"].sort()
    season_entry["last_applied"] = datetime.now().isoformat()
    return ledger

# Position eligibility for each starter slot (mirrors lineup_optimizer.py)
SLOT_POSITIONS = {
    "PG": ["PG"],
    "SG": ["SG"],
    "G": ["PG", "SG"],
    "SF": ["SF"],
    "PF": ["PF"],
    "F": ["SF", "PF"],
    "C": ["C"],
    "UTIL": ["PG", "SG", "SF", "PF", "C"],
}


def _can_fill_slot(positions: list[str], slot: str) -> bool:
    """Check if a player's positions allow them to fill a given starter slot."""
    eligible = SLOT_POSITIONS.get(slot, [])
    return any(pos in eligible for pos in positions)


def _parse_positions(pos_str) -> list[str]:
    """Parse comma-separated position string into a list."""
    if not pos_str or pd.isna(pos_str):
        return []
    return [p.strip() for p in str(pos_str).split(",")]


def normalize_manager(s):
    return str(s).strip()


def compute_weekly_from_playerlog(plog: pd.DataFrame, season: str, week: int):
    """Return per-manager weekly points and healthy starter games."""
    plog = plog.copy()
    plog.columns = [c.strip() for c in plog.columns]

    # Filter to this season + week
    plog = plog[(plog["season_year"] == season) & (plog["week"] == week)].copy()

    if plog.empty:
        return {}, {}

    plog["manager"] = plog["manager"].map(normalize_manager)

    # Ensure boolean
    plog["is_injured"] = plog["is_injured"].fillna(False).astype(bool)
    plog["started"] = plog["started"].fillna(False).astype(bool)

    healthy_starters = plog[(plog["started"]) & (~plog["is_injured"])].copy()
    if healthy_starters.empty:
        return {}, {}

    weekly_points = (
        healthy_starters.groupby("manager")["fantasy_points"].sum().to_dict()
    )
    healthy_games = healthy_starters.groupby("manager").size().to_dict()

    return weekly_points, healthy_games


def compute_weekly_scheduled_from_lineups(lineups: pd.DataFrame, season: str, week: int):
    """Return per-manager weekly scheduled games.

    Definition of scheduled game:
    - Row is in this season & week
    - slot is NOT IL (IL+ is included)
    - nba_opponent is non-blank (non-null, non-empty)
    - fantasy_points has a value (not NaN)
    """
    lineups = lineups.copy()
    lineups.columns = [c.strip() for c in lineups.columns]

    # Filter to this season + week ONLY
    lineups = lineups[
        (lineups["season_year"] == season) &
        (lineups["week"] == week)
    ].copy()
    if lineups.empty:
        return {}

    lineups["manager"] = lineups["manager"].map(normalize_manager)
    lineups["slot"] = lineups["slot"].astype(str).str.upper()
    
    # Ensure fantasy_points is numeric
    lineups["fantasy_points"] = pd.to_numeric(
        lineups["fantasy_points"], errors="coerce"
    )

    # Non-IL only (IL+ is included in scheduled)
    mask_non_il = ~lineups["slot"].isin(["IL"])

    # nba_opponent must be non-null and non-empty
    mask_has_opp = lineups["nba_opponent"].notna() & (
        lineups["nba_opponent"].astype(str).str.strip() != ""
    )
    
    # fantasy_points must have a value (not NaN)
    mask_has_fp = lineups["fantasy_points"].notna()

    scheduled_df = lineups[mask_non_il & mask_has_opp & mask_has_fp]
    if scheduled_df.empty:
        return {}

    scheduled = scheduled_df.groupby("manager").size().to_dict()
    return scheduled


def compute_weekly_games_lost_from_lineups(lineups: pd.DataFrame, season: str, week: int):
    """Return per-manager weekly games lost to injury.

    Definition of game lost to injury:
    - Row is in this season & week
    - slot is NOT IL (IL+ is included)
    - nba_opponent is non-blank (non-null, non-empty)
    - fantasy_points = 0.0
    """
    lineups = lineups.copy()
    lineups.columns = [c.strip() for c in lineups.columns]

    lineups = lineups[
        (lineups["season_year"] == season) &
        (lineups["week"] == week)
    ].copy()
    if lineups.empty:
        return {}

    lineups["manager"] = lineups["manager"].map(normalize_manager)
    lineups["slot"] = lineups["slot"].astype(str).str.upper()
    lineups["fantasy_points"] = pd.to_numeric(
        lineups["fantasy_points"], errors="coerce"
    )

    mask = (
        ~lineups["slot"].isin(["IL"]) &
        lineups["nba_opponent"].notna() &
        (lineups["nba_opponent"].astype(str).str.strip() != "") &
        (lineups["fantasy_points"] == 0.0)
    )

    lost_df = lineups[mask]
    if lost_df.empty:
        return {}

    return lost_df.groupby("manager").size().to_dict()


def compute_weekly_bench_games_from_lineups(lineups: pd.DataFrame, season: str, week: int):
    """Return per-manager weekly games left on bench.

    Definition of game left on bench:
    - Row is in this season & week
    - slot is BN or IL+
    - fantasy_points > 0
    """
    lineups = lineups.copy()
    lineups.columns = [c.strip() for c in lineups.columns]

    lineups = lineups[
        (lineups["season_year"] == season) &
        (lineups["week"] == week)
    ].copy()
    if lineups.empty:
        return {}

    lineups["manager"] = lineups["manager"].map(normalize_manager)
    lineups["slot"] = lineups["slot"].astype(str).str.upper()
    lineups["fantasy_points"] = pd.to_numeric(
        lineups["fantasy_points"], errors="coerce"
    )

    mask = (
        lineups["slot"].isin(["BN", "IL+"]) &
        (lineups["fantasy_points"] > 0)
    )

    bench_df = lineups[mask]
    if bench_df.empty:
        return {}

    return bench_df.groupby("manager").size().to_dict()


def compute_weekly_blunders_from_lineups(lineups: pd.DataFrame, playerlog: pd.DataFrame, season: str, week: int):
    """Return per-manager weekly blunder count.

    A blunder is a bench player who played (FP > 0) when a starter slot was
    available -- either empty or occupied by a player who didn't play that day.
    This requires cross-referencing LINEUPS (for slot assignments) with
    PLAYERLOG (for who actually played) on a per-day basis.

    Uses greedy matching: highest-FP bench players are assigned to available
    slots first to avoid double-counting when multiple bench players could
    fill the same slot.
    """
    lineups = lineups.copy()
    lineups.columns = [c.strip() for c in lineups.columns]
    playerlog = playerlog.copy()
    playerlog.columns = [c.strip() for c in playerlog.columns]

    lineups = lineups[
        (lineups["season_year"] == season) & (lineups["week"] == week)
    ].copy()
    playerlog = playerlog[
        (playerlog["season_year"] == season) & (playerlog["week"] == week)
    ].copy()

    if lineups.empty:
        return {}

    lineups["manager"] = lineups["manager"].map(normalize_manager)
    lineups["slot"] = lineups["slot"].astype(str).str.upper()
    lineups["fantasy_points"] = pd.to_numeric(lineups["fantasy_points"], errors="coerce")

    playerlog["manager"] = playerlog["manager"].map(normalize_manager)
    playerlog["is_injured"] = playerlog["is_injured"].fillna(False).astype(bool)
    playerlog["fantasy_points"] = pd.to_numeric(playerlog["fantasy_points"], errors="coerce")

    managers = lineups["manager"].unique()
    blunder_counts = {}

    for manager in managers:
        mgr_lineups = lineups[lineups["manager"] == manager]
        mgr_plog = playerlog[playerlog["manager"] == manager]
        mgr_blunders = 0

        for day in mgr_lineups["date"].unique():
            lu_day = mgr_lineups[mgr_lineups["date"] == day]
            pl_day = mgr_plog[mgr_plog["date"] == day]

            # Build FP lookup from playerlog (who actually played)
            fp_lookup = {}
            for _, row in pl_day.iterrows():
                if not row.get("is_injured", False):
                    fp_lookup[row["player_name"]] = float(row.get("fantasy_points", 0))

            bench_played = []
            dnp_starters = []

            for _, row in lu_day.iterrows():
                slot = row["slot"]
                player = row["player_name"]
                positions = _parse_positions(row.get("positions", ""))
                fp = fp_lookup.get(player)

                if slot in ("IL", "IL+"):
                    continue

                if slot == "BN":
                    if fp is not None and fp > 0:
                        bench_played.append({"positions": positions, "fp": fp})
                else:
                    if fp is None:
                        # Starter slot with no game played -> available
                        dnp_starters.append({"slot": slot, "positions": positions})

            # Greedy match: highest FP bench player first
            available = list(dnp_starters)
            for bp in sorted(bench_played, key=lambda p: p["fp"], reverse=True):
                for i, dnp in enumerate(available):
                    if _can_fill_slot(bp["positions"], dnp["slot"]):
                        mgr_blunders += 1
                        available.pop(i)
                        break

        if mgr_blunders > 0:
            blunder_counts[manager] = mgr_blunders

    return blunder_counts


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--season",
        default=CURRENT_SEASON_LONG,
        help="Season label as stored in season_year column (default: 2025-2026).",
    )
    parser.add_argument(
        "--week",
        type=int,
        required=True,
        help="Fantasy week number to add into LEAGUEHISTORY.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Re-apply a week even if the ledger says it has already been applied. "
             "Use only after a correction (otherwise this causes double-counting).",
    )
    args = parser.parse_args()

    base = Path(".")

    # --- Applied-weeks ledger guard (prevents double-counting on re-run) ---
    ledger = _load_applied_weeks_ledger(base, args.season)
    if _is_week_already_applied(ledger, args.season, args.week) and not args.force:
        print(
            f"WARNING: Week {args.week} has already been applied to "
            f"LEAGUEHISTORY.xlsx for season {args.season}. Skipping to prevent "
            f"double-counting."
        )
        print(
            "To force re-application (e.g., after a correction), use --force."
        )
        return
    if _is_week_already_applied(ledger, args.season, args.week) and args.force:
        print(
            f"--force: Week {args.week} previously applied; re-accumulating "
            f"per user override."
        )

    # Load source data
    plog = pd.read_excel(base / PLAYERLOG_XLSX_PATH)
    lineups = pd.read_excel(base / LINEUPS_XLSX_PATH)

    weekly_points, healthy_games = compute_weekly_from_playerlog(
        plog, args.season, args.week
    )
    scheduled_games = compute_weekly_scheduled_from_lineups(
        lineups, args.season, args.week
    )
    games_lost = compute_weekly_games_lost_from_lineups(
        lineups, args.season, args.week
    )
    bench_games = compute_weekly_bench_games_from_lineups(
        lineups, args.season, args.week
    )
    blunders = compute_weekly_blunders_from_lineups(
        lineups, plog, args.season, args.week
    )

    print(f"Week {args.week} summary (season {args.season}):")
    managers = sorted(
        set(list(weekly_points.keys()) + list(healthy_games.keys()) + list(scheduled_games.keys()))
    )
    for m in managers:
        print(
            f"  {m}: "
            f"{weekly_points.get(m, 0):.1f} FP, "
            f"{scheduled_games.get(m, 0)} scheduled, "
            f"{healthy_games.get(m, 0)} healthy, "
            f"{games_lost.get(m, 0)} lost, "
            f"{bench_games.get(m, 0)} bench, "
            f"{blunders.get(m, 0)} blunders"
        )

    # Load LEAGUEHISTORY with openpyxl to preserve formulas
    wb = load_workbook(base / LEAGUEHISTORY_XLSX_PATH, data_only=False)
    ws = wb.active  # assume first sheet

    # Header / column indices
    header = [cell.value for cell in ws[1]]
    col_idx = {name: i + 1 for i, name in enumerate(header)}

    required_cols = [
        "manager_name",
        "total_points_scored_to_date",
        "total_points_current_season",
        "total_scheduled_games_current_season",
        "total_healthy_starter_games_current_season",
        "total_games_lost_current_season",
        "total_games_left_on_bench_current_season",
        "total_blunders_current_season",
    ]
    for col in required_cols:
        if col not in col_idx:
            raise ValueError(f"LEAGUEHISTORY.xlsx is missing required column: {col}")

    def get_number(row, col_name):
        c = col_idx[col_name]
        val = ws.cell(row=row, column=c).value
        if val is None or val == "":
            return 0.0
        try:
            return float(val)
        except Exception:
            return 0.0

    # Update each manager row
    for r in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=r, column=col_idx["manager_name"])
        if not name_cell.value:
            continue
        manager = normalize_manager(name_cell.value)

        add_fp = weekly_points.get(manager, 0.0)
        add_sched = scheduled_games.get(manager, 0)
        add_healthy = healthy_games.get(manager, 0)
        add_lost = games_lost.get(manager, 0)
        add_bench = bench_games.get(manager, 0)
        add_blunders = blunders.get(manager, 0)

        if add_fp == 0 and add_sched == 0 and add_healthy == 0:
            continue

        cur_total_pts = get_number(r, "total_points_scored_to_date")
        ws.cell(row=r, column=col_idx["total_points_scored_to_date"], value=cur_total_pts + add_fp)

        cur_season_pts = get_number(r, "total_points_current_season")
        ws.cell(row=r, column=col_idx["total_points_current_season"], value=cur_season_pts + add_fp)

        cur_sched = get_number(r, "total_scheduled_games_current_season")
        ws.cell(row=r, column=col_idx["total_scheduled_games_current_season"], value=cur_sched + add_sched)

        cur_healthy = get_number(r, "total_healthy_starter_games_current_season")
        ws.cell(row=r, column=col_idx["total_healthy_starter_games_current_season"], value=cur_healthy + add_healthy)

        cur_lost = get_number(r, "total_games_lost_current_season")
        ws.cell(row=r, column=col_idx["total_games_lost_current_season"], value=cur_lost + add_lost)

        cur_bench = get_number(r, "total_games_left_on_bench_current_season")
        ws.cell(row=r, column=col_idx["total_games_left_on_bench_current_season"], value=cur_bench + add_bench)

        cur_blunders = get_number(r, "total_blunders_current_season")
        ws.cell(row=r, column=col_idx["total_blunders_current_season"], value=cur_blunders + add_blunders)

    wb.save(base / LEAGUEHISTORY_XLSX_PATH)
    print("LEAGUEHISTORY.xlsx updated.")

    ledger = _record_week_applied(ledger, args.season, args.week)
    _save_applied_weeks_ledger(base, ledger)
    print(f"Recorded week {args.week} ({args.season}) in {APPLIED_WEEKS_LEDGER_PATH}.")


if __name__ == "__main__":
    main()
